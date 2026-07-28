"""
Batch processor for approval instances.

Handles batch downloading of attachments from multiple approval instances.
"""

import json
import logging
import re
import platform
import subprocess
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
    XDRPositiveSize2D,
    pixels_to_EMU,
)
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.feishu_api import download_file, extract_attachments, get_instance_detail, parse_form

logger = logging.getLogger(__name__)

# Platform detection for print functions
HAS_WIN32COM = False
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    pass


def _try_com_progid(progid: str, file_path: Path,
                    printer_name: Optional[str] = None,
                    max_retries: int = 3) -> Tuple[bool, str]:
    """Try printing via a specific COM ProgID (e.g. ``KET.Application``).

    Retries up to *max_retries* times on failure.
    Returns ``(success, error_message)``.
    """
    import time
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    for attempt in range(1, max_retries + 1):
        app = None
        wb = None
        try:
            app = win32com.client.DispatchEx(progid)
            app.Visible = False
            app.DisplayAlerts = False

            wb = app.Workbooks.Open(str(file_path.resolve()))
            ws = wb.ActiveSheet
            ws.PageSetup.Orientation = 2          # xlLandscape
            ws.PageSetup.PaperSize = 9            # xlPaperA4
            ws.PageSetup.FitToPagesWide = 1       # all columns fit to 1 page wide
            ws.PageSetup.FitToPagesTall = 0       # 0=unlimited rows
            # PrintTitleRows intentionally NOT set here — it is managed by
            # payroll_merger.print_file() which knows whether the workbook
            # is a merged "合集" (embedded headers) or a raw signed file.
            if printer_name:
                wb.PrintOut(ActivePrinter=printer_name)
            else:
                wb.PrintOut()
            wb.Close(SaveChanges=False)
            app.Quit()
            return True, ""
        except Exception as e:
            import traceback as _tb
            detail = f"[{progid}] attempt {attempt}/{max_retries}: {type(e).__name__}: {e}"
            logger.warning(f"[PRINT] {detail}")
            if attempt >= max_retries:
                logger.warning(f"[PRINT] traceback: {_tb.format_exc()[:500]}")
            if attempt < max_retries:
                time.sleep(5)
        finally:
            try:
                if wb:
                    wb.Close(SaveChanges=False)
                if app:
                    app.Quit()
            except Exception:
                pass
    return False, detail  # last error detail


def _print_with_com(file_path: Path, printer_name: Optional[str] = None) -> Tuple[bool, str]:
    """Print using WPS/Excel COM (Windows only).

    Tries ``KET.Application`` (WPS) first, then ``Excel.Application``
    (Microsoft Office), each with retries.  Returns ``(success, error_message)``.
    """
    if platform.system() != "Windows":
        return False, "非 Windows 环境"
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return False, "win32com 未安装"

    # WPS COM first → matches the working merge_print path
    ok, err = _try_com_progid("KET.Application", file_path, printer_name)
    if ok:
        return True, ""

    # Fall back to Microsoft Excel COM
    ok, err2 = _try_com_progid("Excel.Application", file_path, printer_name)
    if ok:
        return True, ""
    return False, f"WPS: {err} | Excel: {err2}"


def _print_with_libreoffice(file_path: Path, printer_name: Optional[str] = None) -> Tuple[bool, str]:
    """Print using LibreOffice (cross-platform fallback).

    Returns ``(success, error_message)``.
    """
    try:
        if file_path.suffix.lower() in (".xlsx", ".xls", ".docx", ".doc"):
            cmd = [
                "soffice",
                "--headless",
                "-p",
                str(file_path),
            ]
            if printer_name:
                cmd.extend(["--printer", printer_name])
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=60,
            )
            ok = result.returncode == 0
            return ok, "" if ok else f"soffice exit code {result.returncode}: {result.stderr[:200]}"
        return False, f"不支持的文件格式: {file_path.suffix}"
    except FileNotFoundError:
        msg = "LibreOffice 未安装（soffice 命令未找到）"
        logger.warning(f"[PRINT] {msg}")
        return False, msg


def print_file(file_path: Path, printer_name: Optional[str] = None) -> Tuple[bool, str]:
    """Print file. Windows uses WPS/Excel COM, Linux uses LibreOffice.

    Returns ``(success, error_message)``.
    """
    if platform.system() == "Windows":
        ok, err = _print_with_com(file_path, printer_name)
        if ok:
            return True, ""
        logger.warning(f"[PRINT] COM打印失败 ({err})，尝试LibreOffice...")
        return _print_with_libreoffice(file_path, printer_name)
    else:
        return _print_with_libreoffice(file_path, printer_name)


_PAYROLL_CONFIG_PATH = Path(__file__).parent / "payroll_sheet_config.json"
_ROLE_MAPPING_PATH = Path(__file__).parent / "role_mapping.json"


def _load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _get_mandatory_roles() -> set:
    config = _load_json(_PAYROLL_CONFIG_PATH)
    mandatory = config.get("sheet_filter", {}).get("signatures", {}).get("mandatory", {})
    return {k for k in mandatory.keys() if k != "description"}


def _get_role_mapping(path: Optional[Path] = None) -> Dict[str, str]:
    mapping_path = path or _ROLE_MAPPING_PATH
    data = _load_json(mapping_path)
    return {k: v for k, v in data.items() if not k.startswith("_")}


def is_ready_for_print(details: dict) -> bool:
    """
    Check if the approval is ready for the cashier to process.
    All mandatory signature roles must be approved, and the current
    pending task node must be \"出纳办理\" (cashier step).
    """
    mandatory_roles = _get_mandatory_roles()
    if not mandatory_roles:
        return False

    role_mapping = _get_role_mapping()
    approved_roles: set = set()

    for approver in details.get("approver_list", []):
        if approver.get("status") == "APPROVED":
            role = role_mapping.get(approver.get("approver_name"))
            if role:
                approved_roles.add(role)

    for task in details.get("task_list", []):
        if task.get("status") in ("APPROVED", "DONE"):
            role = role_mapping.get(task.get("node_name", ""))
            if role:
                approved_roles.add(role)

    if not mandatory_roles.issubset(approved_roles):
        return False

    for task in details.get("task_list", []):
        if task.get("status") == "PENDING" and task.get("node_name") == "出纳办理":
            return True
    return False


def is_approval_passed(details: dict) -> bool:
    """
    Check if the overall approval status is APPROVED.

    Args:
        details: Instance detail dict.

    Returns:
        True if details["status"] == "APPROVED".
    """
    return details.get("status") == "APPROVED"


def get_approvers_with_roles(details: dict, role_mapping_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """
    Extract approvers with their mapped roles from the approver_list.

    Args:
        details: Instance detail dict containing "approver_list".
        role_mapping_path: Optional path to a custom role_mapping.json file.

    Returns:
        List of dicts with keys "approver_name", "role", "status".
        If no role mapping is found for an approver_name, "role" is None.
    """
    role_mapping = _get_role_mapping(role_mapping_path)
    result: List[Dict[str, Any]] = []

    for approver in details.get("approver_list", []):
        name = approver.get("approver_name")
        role = role_mapping.get(name) if name else None
        result.append({
            "approver_name": name,
            "role": role,
            "status": approver.get("status"),
        })

    for task in details.get("task_list", []):
        t_status = task.get("status", "")
        if t_status not in ("APPROVED", "DONE"):
            continue
        uid = task.get("user_id", "")
        node = task.get("node_name", "")
        role = role_mapping.get(node)
        if not role:
            continue
        result.append({
            "approver_name": uid,
            "role": role,
            "status": "APPROVED",
        })

    return result


_PAYROLL_CONFIG = None


def reload_payroll_config():
    """Force reload payroll config from disk on next get_payroll_config() call."""
    global _PAYROLL_CONFIG
    _PAYROLL_CONFIG = None


def get_payroll_config() -> dict:
    """Load payroll sheet detection rules from config file."""
    global _PAYROLL_CONFIG
    if _PAYROLL_CONFIG is None:
        if _PAYROLL_CONFIG_PATH.exists():
            try:
                with open(_PAYROLL_CONFIG_PATH, "r", encoding="utf-8") as f:
                    _PAYROLL_CONFIG = json.load(f)
            except (json.JSONDecodeError, OSError) as e:
                logger.warning(f"加载工资表配置失败: {e}，使用内置默认值")
        if _PAYROLL_CONFIG is None:
            _PAYROLL_CONFIG = {
                "sheet_filter": {
                    "row1_title": {"required_keyword": "工资发放表"},
                    "row2_org": {"required_keyword": "单位名称"},
                    "row3_headers": {"required": ["转账合计", "应发工资", "实发工资", "实发合计"]},
                    "signatures": {
                        "mandatory": {
                            "总经理签字": ["总经理签字"],
                        },
                        "optional": {
                            "分管领导审核": ["分管领导审核"],
                            "财务审核": ["财务审核"],
                        },
                    },
                }
            }
        if "text_normalization" not in _PAYROLL_CONFIG:
            _PAYROLL_CONFIG["text_normalization"] = {
                "rules": [
                    {"source": "部长、分管副总签字", "target": "分管领导审核"},
                    {"source": "部长签字", "target": "分管领导审核"},
                ]
            }
    return _PAYROLL_CONFIG


def _get_signature_keywords(cfg: dict) -> set:
    """Extract all signature-related keywords from config.

    Collects keys from mandatory/optional signature configs and source
    fields from text normalization rules into a deduplicated set.
    """
    keywords: set[str] = set()

    mandatory = cfg.get("sheet_filter", {}).get("signatures", {}).get("mandatory", {})
    keywords.update(mandatory.keys())

    optional = cfg.get("sheet_filter", {}).get("signatures", {}).get("optional", {})
    keywords.update(optional.keys())

    rules = cfg.get("text_normalization", {}).get("rules", [])
    for rule in rules:
        source = rule.get("source")
        if source:
            keywords.add(source)

    return keywords


def _flatten_header_merges(ws, header_rows: int = 5) -> None:
    """Unmerge all header-area merges and fill every cell with the anchor value.

    Must be called **before** ``delete_cols``, while merged range addresses
    still match the actual cell positions.  After flattening, column deletion
    shifts ordinary cell values correctly — there are no stale merge ranges
    to cause ``MergedCell`` errors.

    After flattening, rows 4–*header_rows* (sub‑header area) are cleaned so
    cells whose value equals the row‑3 column header (artefacts from vertical
    merges like ``D3:D5``) are cleared back to ``None``.  Legitimate sub‑header
    values (e.g. ``Q4`` = ``"养老"`` under ``Q3`` = ``"扣款明细"``) are kept.
    """
    for mr in list(ws.merged_cells.ranges):
        mc, mr_min, mxc, mr_max = mr.bounds
        if mr_min > header_rows:
            continue
        anchor_val = ws.cell(row=mr_min, column=mc).value
        # Snapshot original formatting before unmerge
        src = ws.cell(row=mr_min, column=mc)
        fmt = {
            'font': copy(src.font),
            'alignment': copy(src.alignment),
            'border': copy(src.border),
        }
        try:
            ws.unmerge_cells(str(mr))
        except KeyError:
            pass
        for r in range(mr_min, mr_max + 1):
            for c in range(mc, mxc + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    try:
                        cell.value = anchor_val
                    except AttributeError:
                        pass
                # Restore formatting to all cells that were MergedCell
                cell.font = copy(fmt['font'])
                cell.alignment = copy(fmt['alignment'])
                cell.border = copy(fmt['border'])

    # Clear duplicate header labels from vertical‑merge rows (4..header_rows)
    for row in range(header_rows, 3, -1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            above = ws.cell(row=3, column=col).value
            if cell.value is not None and cell.value == above:
                cell.value = None


def _rebuild_header_merges(ws, header_rows: int = 5) -> None:
    """Re‑build header merges from flat cell values (rows 1‑*header_rows*).

    Follows the merge algorithm from ``virtual-header-generation.md``:
    1. Row‑4 horizontal merge  — within each row‑3 group, merge adjacent
       row‑4 cells with the same value.
    2. Vertical merges          — three cases (A: 3‑row, B: 2‑row, C: 2‑row).
    3. Catch‑all 3‑row merge    — rows 4‑5 both empty, merge rows 3‑5.
    4. Row‑3 horizontal merge   — last, so row‑4 grouping is unaffected.

    Row‑5 horizontal merging is **not** performed — original sub‑sub‑headers
    (e.g. "单位"/"个人") are always individual cells.

    Precondition: ``_flatten_header_merges`` must have been called before
    column deletion so that every header cell holds its own value.
    """
    H = header_rows
    max_col = ws.max_column
    _cv = lambda r, c: ws.cell(row=r, column=c).value
    hmerged = set()  # (row, col) of already-merged anchor cells

    # ── 1. Row‑4 horizontal merge (within row‑3 groups) ──
    vi = 1
    while vi <= max_col:
        r3v = _cv(3, vi)
        # find row‑3 group [vi, vj)
        vj = vi + 1
        while vj <= max_col and _cv(3, vj) == r3v:
            vj += 1
        # in this group, merge adjacent row‑4 same values
        vk = vi
        while vk < vj:
            r4v = _cv(4, vk)
            if r4v in (None, ''):
                vk += 1
                continue
            vl = vk + 1
            while vl < vj and _cv(4, vl) == r4v:
                vl += 1
            if vl - 1 > vk:
                try:
                    ws.merge_cells(start_row=4, start_column=vk,
                                   end_row=4, end_column=vl - 1)
                except Exception:
                    pass
                hmerged.update((4, c) for c in range(vk, vl))
            vk = vl
        vi = vj

    # ── 2. Vertical merges (3 cases) ──
    for vi in range(1, max_col + 1):
        if (4, vi) in hmerged or (5, vi) in hmerged:
            continue
        r3v, r4v, r5v = _cv(3, vi), _cv(4, vi), _cv(5, vi)
        if r4v and r4v == r5v:
            if r3v and r3v == r4v:
                # Case A — merge 3 rows
                try:
                    ws.merge_cells(start_row=3, start_column=vi,
                                   end_row=H, end_column=vi)
                except Exception:
                    pass
                hmerged.update((r, vi) for r in range(3, H + 1))
            else:
                # Case B — merge 2 rows (4-5)
                try:
                    ws.merge_cells(start_row=4, start_column=vi,
                                   end_row=H, end_column=vi)
                except Exception:
                    pass
                hmerged.update((r, vi) for r in range(4, H + 1))
        elif r4v and not r5v:
            # Case C — merge 2 rows (4-5)
            try:
                ws.merge_cells(start_row=4, start_column=vi,
                               end_row=H, end_column=vi)
            except Exception:
                pass
            hmerged.update((r, vi) for r in range(4, H + 1))

    # ── 3. Catch‑all 3‑row merge ──
    for vi in range(1, max_col + 1):
        if any((r, vi) in hmerged for r in range(3, H + 1)):
            continue
        r4v, r5v = _cv(4, vi), _cv(5, vi)
        if not r4v and not r5v:
            try:
                ws.merge_cells(start_row=3, start_column=vi,
                               end_row=H, end_column=vi)
            except Exception:
                pass

    # ── 4. Row‑3 horizontal merge (last) ──
    vi = 1
    while vi <= max_col:
        r3v = _cv(3, vi)
        if r3v in (None, ''):
            vi += 1
            continue
        vj = vi + 1
        while vj <= max_col and _cv(3, vj) == r3v:
            vj += 1
        if vj - 1 > vi:
            try:
                ws.merge_cells(start_row=3, start_column=vi,
                               end_row=3, end_column=vj - 1)
            except Exception:
                pass
        vi = vj

    # ── 5. Merge row 1 (title) and row 2 (org name + date) — keep existing
    #     horizontal-merge logic only for rows 1-2 (simple single-row merges) ──
    for row in [1, 2]:
        start = None
        prev_val = None
        for col in range(1, ws.max_column + 1):
            v = _cv(row, col)
            if v is not None and v != '' and v == prev_val:
                continue
            if start is not None and col - 1 > start:
                try:
                    ws.merge_cells(
                        start_row=row, start_column=start,
                        end_row=row, end_column=col - 1
                    )
                except Exception:
                    pass
            start = col if (v is not None and v != '') else None
            prev_val = v
        if start is not None and ws.max_column > start:
            try:
                ws.merge_cells(
                    start_row=row, start_column=start,
                    end_row=row, end_column=ws.max_column
                )
            except Exception:
                pass


def _delete_cols_with_merge(ws, col, amount=1, saved_all=None):
    """Delete *amount* columns, then restore every merged range correctly.

    openpyxl's ``delete_cols`` does **not** adjust existing merged ranges
    — they keep their original column-letter bounds, leaving stale
    ``MergedCell`` objects behind.

    This function takes a snapshot of **all** current merged ranges,
    deletes the columns, then rebuilds each range with the deleted span
    removed (or shifted left for ranges entirely to the right).

    If *saved_all* is provided (a full snapshot obtained before any
    explicit ``unmerge_cells`` call), it is used instead of taking a new
    snapshot — needed when the caller has already unmerged some ranges
    before deletion.
    """
    if saved_all is None:
        saved_all = list(ws.merged_cells.ranges)

    ws.delete_cols(col, amount)

    for mr in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(mr))
        except Exception:
            pass

    for mr in saved_all:
        mc, r0, mxc, rmax = mr.bounds
        if mxc < col:               # entirely left — unchanged
            new_mc, new_mxc = mc, mxc
        elif mc >= col + amount:    # entirely right — shift left
            new_mc, new_mxc = mc - amount, mxc - amount
        else:                       # intersects — shrink + shift
            new_mc = mc if mc < col else max(col, mc - amount)
            new_mxc = mxc if mxc < col else mxc - amount
        if new_mc <= new_mxc:
            try:
                ws.merge_cells(start_row=r0, start_column=new_mc,
                               end_row=rmax, end_column=new_mxc)
            except Exception:
                pass


def _remove_empty_columns(ws, cfg, formula_values: Optional[Dict] = None) -> None:
    """Delete columns with no real data, preserving signature keywords.

    Scans columns right-to-left. **Header rows (1‑3) are ignored** when
    determining whether a column is empty — only the data area (row 4+)
    is considered.

    A column is deleted when **all** its data-area values are:

    * ``None``, **or**
    * signature keywords, **or**
    * formulas (strings starting with ``=``).

    Keywords and **computed formula values** are copied to the nearest
    non‑empty column on the right first (or appended at the end) so they
    are never lost.  *formula_values* maps ``(row, col) → computed_value``
    (loaded with ``data_only=True``); when present, the computed number is
    written instead of the raw formula string.
    """
    keywords = _get_signature_keywords(cfg)
    removed = []
    moved = []

    DATA_START = 4  # skip title/unit/column-header rows

    def _is_formula(val: str) -> bool:
        return val.startswith("=")

    def _is_removable(val: str) -> bool:
        return any(kw in val for kw in keywords) or _is_formula(val)

    for col in range(ws.max_column, 0, -1):
        non_empty = {}
        for row in range(DATA_START, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            # xlsx 空单元格可能是 '' 而非 None，两者都跳过
            if v not in (None, ''):
                non_empty[row] = str(v)

        if not non_empty:
            _delete_cols_with_merge(ws, col)
            removed.append(col)
            continue

        if not all(_is_removable(v) for v in non_empty.values()):
            continue

        target = None
        for rc in range(col + 1, ws.max_column + 1):
            for r in range(DATA_START, ws.max_row + 1):
                v = ws.cell(row=r, column=rc).value
                if v not in (None, '') and not _is_removable(str(v)):
                    target = rc
                    break
            if target is not None:
                break

        if target is None:
            target = ws.max_column + 1

        # Snapshot ALL merged ranges BEFORE any unmerge so that even
        # ranges we're about to unmerge are captured and restored
        # correctly after column deletion.
        merge_snapshot = list(ws.merged_cells.ranges)

        for row, val in non_empty.items():
            # Evaluate formulas to static values to avoid broken references
            if _is_formula(val):
                computed = formula_values.get((row, col)) if formula_values else None
                if computed is not None:
                    val = computed

            cell = ws.cell(row=row, column=target)
            src_cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in list(ws.merged_cells.ranges):
                    mc_min, mc_min_row, mc_max, mc_max_row = mr.bounds
                    if mc_min_row <= row <= mc_max_row and mc_min <= target <= mc_max:
                        try:
                            ws.unmerge_cells(str(mr))
                        except KeyError:
                            pass
                        break
            tgt = ws.cell(row=row, column=target)
            tgt.value = val
            # Copy formatting from the source (being deleted) to the target
            if tgt is not src_cell:
                tgt.font = copy(src_cell.font)
                tgt.alignment = copy(src_cell.alignment)
                tgt.border = copy(src_cell.border)
                tgt.number_format = copy(src_cell.number_format)
            moved.append((col, target, val))

        _delete_cols_with_merge(ws, col, saved_all=merge_snapshot)
        removed.append(col)

    if removed:
        logger.info(f"[CLEANUP] Removed empty columns: {removed}")
    if moved:
        for s, d, v in moved:
            logger.info(f"[CLEANUP] Moved '{v}' from col {s} to col {d}")


def _remove_force_delete_columns(ws, cfg, formula_values: Optional[Dict] = None) -> None:
    """Delete columns that match configured header names, even if they have data.

    Some columns (e.g. ``岗位``) contain real values per employee but the
    finance department does not need them in the printed output.  This
    function removes them in the same fashion as empty‑column deletion,
    preserving only formula values (e.g. per‑employee subtotals) by moving
    them to the nearest non‑deleted column to the right.

    Must be called after ``_flatten_header_merges`` and before
    ``_rebuild_header_merges`` so that row‑3 header names are reliable
    and merged ranges are not yet rebuilt.
    """
    force_cfg = cfg.get("force_delete_columns", {})
    force_headers = {h.strip() for h in force_cfg.get("columns", [])}
    if not force_headers:
        return

    # Identify columns to delete by matching row‑3 header text
    cols_to_delete = set()
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val and str(val).strip() in force_headers:
            cols_to_delete.add(col)

    if not cols_to_delete:
        return

    logger.info(f"[FORCE-DELETE] Headers to force-delete: {force_headers} "
                f"(cols {sorted(cols_to_delete)})")

    DATA_START = 4

    # Process right‑to‑left so that column indices remain valid
    for col in sorted(cols_to_delete, reverse=True):
        # Collect formula values that need to survive
        formula_vals: Dict[int, Any] = {}
        for row in range(DATA_START, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v not in (None, ''):
                raw = str(v)
                if raw.startswith('='):
                    computed = formula_values.get((row, col)) if formula_values else None
                    if computed is not None:
                        formula_vals[row] = computed

        # Move formula values to a safe column (skip other force‑deleted cols)
        if formula_vals:
            target = None
            for tc in range(col + 1, ws.max_column + 1):
                if tc not in cols_to_delete:
                    target = tc
                    break
            if target is None:
                target = ws.max_column + 1

            merge_snapshot = list(ws.merged_cells.ranges)
            for row, val in formula_vals.items():
                cell = ws.cell(row=row, column=target)
                if isinstance(cell, MergedCell):
                    for mr in list(ws.merged_cells.ranges):
                        mc_min, mc_min_row, mc_max, mc_max_row = mr.bounds
                        if mc_min_row <= row <= mc_max_row and mc_min <= target <= mc_max:
                            try:
                                ws.unmerge_cells(str(mr))
                            except KeyError:
                                pass
                            break
                tgt = ws.cell(row=row, column=target)
                tgt.value = val
                src_cell = ws.cell(row=row, column=col)
                if tgt is not src_cell:
                    tgt.font = copy(src_cell.font)
                    tgt.alignment = copy(src_cell.alignment)
                    tgt.border = copy(src_cell.border)
                    tgt.number_format = copy(src_cell.number_format)

        _delete_cols_with_merge(ws, col)

    logger.info(f"[FORCE-DELETE] Deleted columns: {sorted(cols_to_delete)}")


def _apply_normalization_rules(cell_value: str, rules: list) -> str:
    """Apply text normalization rules to a cell value."""
    for rule in rules:
        if rule["source"] in cell_value:
            cell_value = cell_value.replace(rule["source"], rule["target"])
    return cell_value


def is_payroll_sheet(ws, config: Optional[dict] = None) -> bool:
    cfg = config or get_payroll_config()
    sf = cfg["sheet_filter"]

    all_text = ""
    header_text = ""
    for row in range(1, min(ws.max_row + 1, 50)):
        row_text = ""
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                cleaned = str(cell_val).replace("\u3000", "").strip()
                row_text += cleaned
                all_text += cleaned
        if row <= 3:
            header_text += row_text

    for kw in sf.get("exclude_keywords", {}).get("keywords", []):
        if kw in header_text:
            return False

    title_kw = sf.get("title_keyword", {}).get("required", "")
    if title_kw and title_kw not in header_text:
        return False

    for kw in sf.get("required_content", {}).get("required", []):
        if kw not in all_text:
            return False

    mandatory = {
        k: v for k, v in sf["signatures"]["mandatory"].items()
        if k != "description"
    }
    if mandatory:
        found = set()
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = str(ws.cell(row=row, column=col).value or "")
                for role, keywords in mandatory.items():
                    if role not in found:
                        if any(kw in cell_val for kw in keywords):
                            found.add(role)
                if len(found) == len(mandatory):
                    break
            if len(found) == len(mandatory):
                break
        if len(found) < len(mandatory):
            return False

    return True


def _has_unit_name_in_row2(ws) -> bool:
    """Check if row 2 contains a unit / institution name.

    System-generated payroll sheets have a ``名称：XXX`` pattern on row 2
    (e.g. ``A2=单位  B2=名称：供销粮油吉林有限公司（外包）``).  Manual
    sheets use row 2 directly as column headers (``序号``, ``姓名``,
    ``基本工资``…) and should skip column‑deletion operations.
    """
    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row=2, column=col)
        if cell.value:
            text = str(cell.value).strip()
            if "名称：" in text or text.startswith("单位名称"):
                return True
    return False


def _is_cell_in_merged_range(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    return None


def _split_merged_for_text(ws, row, col):
    cfg = get_payroll_config()
    rules = cfg.get("text_normalization", {}).get("rules", [])
    merged = _is_cell_in_merged_range(ws, row, col)
    if not merged:
        cell = ws.cell(row=row, column=col)
        if cell.value:
            cell.value = _apply_normalization_rules(str(cell.value), rules)
        return col + 1

    cell = ws.cell(row=row, column=col)
    text = str(cell.value) if cell.value else ""
    normalized = _apply_normalization_rules(text, rules)

    if normalized != text:
        cell.value = normalized
        needed_cols = 3
    else:
        needed_cols = 2

    total_cols = merged.max_col - merged.min_col + 1
    if needed_cols >= total_cols:
        return merged.max_col + 1

    merged_str = str(merged)
    ws.unmerge_cells(merged_str)

    new_end = merged.min_col + needed_cols - 1
    ws.merge_cells(start_row=row, start_column=merged.min_col,
                   end_row=row, end_column=new_end)

    for c in range(new_end + 1, merged.max_col + 1):
        ws.cell(row=row, column=c).value = None

    return new_end + 1


def find_all_signature_positions(ws, config: Optional[dict] = None) -> Dict[str, Tuple[int, int]]:
    positions = {}
    cfg = config or get_payroll_config()
    sf = cfg["sheet_filter"]["signatures"]

    keyword_groups = []
    for role, keywords in sf.get("mandatory", {}).items():
        keyword_groups.append((keywords, role))
    for role, keywords in sf.get("optional", {}).items():
        keyword_groups.append((keywords, role))

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                value = str(cell.value).strip()
                for texts, role_key in keyword_groups:
                    if role_key not in positions:
                        for text in texts:
                            if text in value:
                                positions[role_key] = (row, col)
                                break
    return positions


# ── Standard filename helpers ──────────────────────────────────────────────

# Keywords that identify a cell value as a unit/institution name.
_UNIT_KEYWORDS = frozenset({
    "大学", "学院", "公司", "集团", "学校", "委员会",
    "研究院", "中心", "局", "处", "厂", "社", "所", "企业",
})


def _is_unit_name(text: str) -> bool:
    if len(text) < 3:
        return False
    for kw in _UNIT_KEYWORDS:
        if len(kw) >= 2:
            if kw in text:
                return True
        else:
            # 单字关键词（局、处、厂、社、所）仅在字符串末尾匹配，
            # 避免"应纳税所得额"（含"所"）等非单位词误判。
            if text.endswith(kw):
                return True
    return False


def _strip_label_prefix(text: str) -> str:
    """Strip a prefixed label such as ``名称：XXX`` → ``XXX``."""
    m = re.match(r'^[^：]+：\s*', text)
    if m:
        return text[m.end():]
    return text


def _extract_unit_name(ws) -> Optional[str]:
    """Try to extract the unit name from a payroll worksheet.

    Resolution order:
      1. If row 2 contains a unit name (detected by :func:`_has_unit_name_in_row2`
         — ``名称：`` or ``单位名称`` pattern), scan row 2 cells and return
         the first match.  This handles system‑generated sheets where row 2
         stores the unit.
      2. Otherwise (manual sheets — row 2 is column headers like 序号/姓名/
         基本工资), skip row 2 and inspect row 1.  If the title matches the
         pattern ``XXX\\d{4}年…``, extract the leading text and check it
         against :func:`_is_unit_name`.
      3. Return ``None`` if no unit name is found.
    """
    if _has_unit_name_in_row2(ws):
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value:
                name = _strip_label_prefix(str(cell.value).strip())
                if _is_unit_name(name):
                    return name

    title = None
    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            title = str(cell.value).strip()
            break

    if title:
        # Pattern: <unit-name><digits>年… — grab everything before the year digits
        m = re.match(r'^(.+?)\d{4}年', title)
        if m:
            candidate = m.group(1).strip()
            if candidate and _is_unit_name(candidate):
                return candidate

    return None


def _extract_year_month(ws) -> Optional[str]:
    """Extract a ``YYYY年MM月`` string from the row-1 title.

    Tries two patterns in order:
      1. ``\\d{4}年\\d{1,2}月``  (e.g. ``2026年07月``)
      2. ``YYYYMM`` — six consecutive digits where the last two are 01‑12

    The month component is always zero-padded to two digits.
    Returns ``None`` when neither pattern is found.
    """
    title = None
    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            title = str(cell.value).strip()
            break

    if title:
        # 1) 年/月 格式
        m = re.search(r'(\d{4})年(\d{1,2})月', title)
        if m:
            year = m.group(1)
            month = m.group(2).zfill(2)
            return f"{year}年{month}月"

        # 2) YYYYMM 连写格式
        for m in re.finditer(r'(?<!\d)(\d{4})(\d{2})(?!\d)', title):
            y, mo = m.group(1), m.group(2)
            if 1 <= int(mo) <= 12:
                return f"{y}年{mo}月"

    return None


def _is_standard_filename(name: str) -> bool:
    """Return True if *name* already contains a unit name + year-month.

    A "standard" payroll filename carries at least:
      1. A recognisable unit/organisation name at the start
      2. A year-month indicator — either ``YYYY年MM月`` or ``YYYYMM``
         (6 consecutive digits, possibly after a short alphanumeric prefix)

    Suffixes such as 工资, 工资表 and 系统 are all accepted — only the
    presence of a unit name **and** a year-month matters.
    """
    stem = Path(name).stem
    if stem.lower().startswith("signed_"):
        stem = stem[7:]

    # 1) YYYY年MM月 格式
    m = re.search(r'(\d{4})年\d{1,2}月', stem)
    if m:
        unit_candidate = stem[: m.start()]
        return bool(unit_candidate and _is_unit_name(unit_candidate))

    # 2) YYYYMM 连写格式
    m = re.search(r'(\d{6})', stem)
    if m:
        unit_candidate = stem[: m.start()]
        return bool(unit_candidate and _is_unit_name(unit_candidate))

    # 3) YYYY.MM 或 YYYY-MM 格式（2026.07 或 2026-07）
    m = re.search(r'(\d{4})[.\-](\d{2})', stem)
    if m:
        unit_candidate = stem[: m.start()]
        return bool(unit_candidate and _is_unit_name(unit_candidate))

    return False


def _build_standard_name(excel_path: Path, ws) -> Optional[str]:
    """Build a standard payroll filename from worksheet metadata.

    Uses :func:`_extract_unit_name` and :func:`_extract_year_month`; returns
    ``None`` when either piece is missing.
    """
    unit = _extract_unit_name(ws)
    ym = _extract_year_month(ws)
    if unit and ym:
        return f"{unit}{ym}工资表.xlsx"
    return None


# ── Legacy helpers ─────────────────────────────────────────────────────────


def _extract_first_row_title(ws) -> Optional[str]:
    """Find the first non-empty cell in row 1 as the table title."""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            return str(cell.value).strip()
    return None


def _build_output_path(excel_path: Path, output_path: Path, ws) -> Path:
    """Build the final signed output path with optional standard renaming.

    Resolution order:
      1. If the source filename does **not** already match the standard
         payroll format, try to build one from worksheet metadata.
      2. Fallback: for raw Feishu exports (``tddd_dialog*``), append the
         row-1 title to the filename so it becomes distinguishable.
      3. Return the original *output_path* unchanged.
    """
    original_name = excel_path.name

    if not _is_standard_filename(original_name):
        standard = _build_standard_name(excel_path, ws)
        if standard:
            return output_path.parent / f"signed_{standard}"

    if original_name.lower().startswith("tddd_dialog"):
        title = _extract_first_row_title(ws)
        if title:
            safe_title = sanitize_dir_name(title)
            new_name = f"signed_{Path(original_name).stem}-{safe_title}.xlsx"
            return output_path.parent / new_name

    return output_path


def _find_total_row(ws) -> int:
    keywords = ["合计", "总计", "合计金额", "合计费用", "合计支付"]
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                # 去空格后匹配，兼容"合 计"等含空格的写法
                val = str(cell.value).replace(" ", "").replace("\u3000", "")
                for kw in keywords:
                    if kw in val and len(val) <= len(kw) + 4:
                        return row
    return 0


def _estimate_col_width(cell_value) -> float:
    """估算单元格内容所需列宽。中文/全角字符计 2，其他计 1。"""
    text = str(cell_value)
    width = 0
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff' or '\uff00' <= ch <= '\uffef':
            width += 2
        else:
            width += 1
    return width


def _calc_data_font_size(ws, col_widths: dict, formula_values: Optional[dict] = None) -> float:
    """
    根据列宽与内容宽度的比值，动态选择一个统一的数据区字号。

    取所有可见列中「最紧」那列的比值，映射到字号：
      col_width / content_width >= 2.0 → 16pt
      col_width / content_width >= 1.4 → 14pt
      其他 → 11pt

    注意：内容宽度从表头行（row 3）开始扫描，排除标题行（row 1）和签名行的长文本干扰。
    """
    sig_keywords = {"总经理签字", "部长签字", "财务审核", "业务审核", "部长、分管副总签字", "分管副总签字"}
    min_ratio = float("inf")
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        if ws.column_dimensions[col_letter].hidden:
            continue
        col_w = col_widths.get(col, 0)
        if col_w <= 0:
            continue
        max_cw = 0
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None:
                # 跳过签名行
                is_sig = False
                for c in range(1, ws.max_column + 1):
                    cv = ws.cell(row=row, column=c).value
                    if cv and any(kw in str(cv) for kw in sig_keywords):
                        is_sig = True
                        break
                if is_sig:
                    continue
                if isinstance(val, str) and val.startswith("="):
                    # 优先用 data_only 缓存值估算，无缓存则跳过
                    if formula_values and (row, col) in formula_values:
                        val = formula_values[(row, col)]
                    else:
                        continue
                w = _estimate_col_width(val)
                if w > max_cw:
                    max_cw = w
        if max_cw > 3:
            ratio = col_w / max_cw
            if ratio < min_ratio:
                min_ratio = ratio

    if min_ratio >= 2.0:
        return 16
    elif min_ratio >= 1.4:
        return 14
    return 11


def _is_sig_keyword_row(ws, row: int, sig_keywords: list) -> bool:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if v and any(kw in str(v) for kw in sig_keywords):
            return True
    return False


def _auto_column_width(ws, cfg=None, formula_values: Optional[dict] = None,
                       min_width: float = 6, max_width: float = 20):
    """
    自适应列宽 + 统一数据区字号，避免打印时 ### 溢出或列过宽导致缩放字太小。

    流程：
    1. 自适应列宽（只扩不缩，以合计行/签名行为数据区下界）
    2. 按各列「列宽/内容宽度」比值动态确定统一字号
    3. 覆盖全表数据行（row 4 起），统一字号
    """
    if cfg is None:
        cfg = get_payroll_config()
    sig_keywords = _get_signature_keywords(cfg)
    total_row = _find_total_row(ws)

    # --- 1. 自适应列宽（取合计行最宽值 + 表头）---
    col_widths = {}
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_content = 0

        # 表头宽度（row 1-3 最宽单元格）
        for row in range(1, min(4, ws.max_row + 1)):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None and not (isinstance(val, str) and val.startswith("=")):
                w = _estimate_col_width(val)
                if w > max_content:
                    max_content = w

        # 合计行宽度（通常该行数字最大）
        if total_row > 0:
            cell = ws.cell(row=total_row, column=col)
            val = cell.value
            if val is not None:
                if isinstance(val, str) and val.startswith("="):
                    # 优先用 data_only 计算值；无缓存时回退到公式文本
                    # （公式文本一定比结果长，安全略宽）
                    if formula_values and (total_row, col) in formula_values:
                        w = _estimate_col_width(formula_values[(total_row, col)])
                    else:
                        w = _estimate_col_width(val)
                else:
                    w = _estimate_col_width(val)
                if w > max_content:
                    max_content = w

        if max_content > 0:
            orig = ws.column_dimensions[col_letter].width or 0
            desired = max(min(max_content + 2, max_width), min_width, orig)
            ws.column_dimensions[col_letter].width = desired
            col_widths[col] = desired

    # --- 2. 动态计算统一字号 ---
    data_font_size = _calc_data_font_size(ws, col_widths, formula_values)

    # --- 3. 统一数据区字号（row 4 起全表覆盖，跳过签名行）---
    for row in range(4, ws.max_row + 1):
        is_sig_row = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=c).value
            if v and any(kw in str(v) for kw in sig_keywords):
                is_sig_row = True
                break
        if is_sig_row:
            continue
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and cell.font.size and cell.font.size != data_font_size:
                cell.font = Font(size=data_font_size, name=cell.font.name)

    # --- 4. 字号调整后复查列宽，避免数字 #### 溢出 + 文本换行 ---
    if data_font_size > 11:
        _SCALE = data_font_size / 11.0
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cur_w = ws.column_dimensions[col_letter].width or 0
            if cur_w <= 0:
                continue
            needed = 0
            has_numeric = False
            has_text_overflow = False
            for row in range(4, ws.max_row + 1):
                if _is_sig_keyword_row(ws, row, sig_keywords):
                    continue
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is None:
                    continue
                if isinstance(val, str) and val.startswith("="):
                    if formula_values and (row, col) in formula_values:
                        val = formula_values[(row, col)]
                    else:
                        continue
                sw = _estimate_col_width(val) * _SCALE
                if sw > needed:
                    needed = sw
                if isinstance(val, (int, float)) or (
                    isinstance(val, str) and re.match(r'^[\d\-,.]', val)
                ):
                    has_numeric = True
                elif sw > cur_w:
                    has_text_overflow = True
            if has_numeric and needed > cur_w:
                ws.column_dimensions[col_letter].width = round(needed + 0.5)
                logger.info(
                    f"[COL] 数值列 {col_letter} 加宽 "
                    f"{cur_w}→{round(needed + 0.5)} "
                    f"(字号 {data_font_size}pt)"
                )
            if has_text_overflow:
                for row in range(4, ws.max_row + 1):
                    if _is_sig_keyword_row(ws, row, sig_keywords):
                        continue
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal if cell.alignment else None,
                            vertical=cell.alignment.vertical if cell.alignment else None,
                            wrap_text=True,
                        )
                logger.info(f"[COL] 文本列 {col_letter} 开启自动换行")


def _hide_columns(ws):
    headers_to_hide = {"部门", "岗位", "职工号"}
    header_row = 3
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        col_letter = get_column_letter(col)
        hidden = False  # reset — after column deletion, original hidden
                       # states no longer correspond to current data
        if cell.value and str(cell.value).strip() in headers_to_hide:
            hidden = True
        ws.column_dimensions[col_letter].hidden = hidden


def _prevent_signature_page_split(ws, positions: Dict[str, Tuple[int, int]]):
    """
    防止签名图片行被分页线切成两半：
    将签名行高度设为与图片匹配（60px ≈ 45pt），确保图片完整容纳在一行内。
    """
    sig_rows = sorted({r for r, _ in positions.values()})
    if not sig_rows:
        return
    for r in sig_rows:
        ws.row_dimensions[r].height = 45
    logger.info(f"[PAGE] 签名行高度调整为45pt ({len(sig_rows)}个签名行)")


def adjust_excel_for_print(ws, cfg=None, formula_values: Optional[dict] = None) -> None:
    """
    调整 Excel 打印设置：横向打印，A4 纸，左边距 2cm，其他边距 1cm，
    所有列缩放到 1 页宽，水平居中。
    在嵌入签名图片前调用。
    """
    if cfg is None:
        cfg = get_payroll_config()
    try:
        ws.page_setup.paperSize = 9          # A4
        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.8
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        ws.print_options.gridLines = False
        logger.info("[PRINT] 已调整: 横向A4, 左2cm其余1cm, 1页宽, fitToPage=True, 网格线关闭")

        _hide_columns(ws)
        _auto_column_width(ws, cfg, formula_values)
    except Exception as e:
        logger.warning(f"[PRINT] 调整打印设置时出错: {e}")


def _convert_xls_to_xlsx_libreoffice(xls_path: Path) -> Optional[Path]:
    """Convert .xls to .xlsx using LibreOffice (cross-platform fallback)."""
    xlsx_path = xls_path.with_suffix(".xlsx")
    try:
        result = subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to", "xlsx",
                str(xls_path.resolve()),
                "--outdir", str(xls_path.parent.resolve()),
            ],
            capture_output=True, text=True, timeout=60,
        )
        if result.returncode == 0 and xlsx_path.exists():
            return xlsx_path
        logger.warning(f"[CONVERT] LibreOffice conversion failed: {result.stderr}")
    except FileNotFoundError:
        logger.warning("[CONVERT] LibreOffice not installed")
    except Exception as e:
        logger.warning(f"[CONVERT] Conversion error: {e}")
    return None


def _convert_xls_to_xlsx_windows(xls_path: Path) -> Optional[Path]:
    """Try WPS/Excel COM first, fallback to LibreOffice."""
    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            xlsx_path = xls_path.with_suffix(".xlsx")
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(xls_path.resolve()))
            wb.SaveAs(str(xlsx_path.resolve()), FileFormat=51)
            wb.Close(SaveChanges=False)
            excel.Quit()
            return xlsx_path
        except Exception as e:
            logger.warning(f"[CONVERT] WPS/Excel COM failed: {e}, trying without FileFormat...")
            try:
                xlsx_path = xls_path.with_suffix(".xlsx")
                wb = excel.Workbooks.Open(str(xls_path.resolve()))
                wb.SaveAs(str(xlsx_path.resolve()))
                wb.Close(SaveChanges=False)
                excel.Quit()
                return xlsx_path
            except Exception as e2:
                logger.warning(f"[CONVERT] COM retry failed: {e2}, falling back to LibreOffice")
                return _convert_xls_to_xlsx_libreoffice(xls_path)
        finally:
            try:
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()
    except ImportError:
        logger.warning("[CONVERT] pywin32 not installed, falling back to LibreOffice")
        return _convert_xls_to_xlsx_libreoffice(xls_path)


def _convert_xls_to_xlsx(xls_path: Path) -> Optional[Path]:
    """Convert .xls to .xlsx: Windows uses WPS/Excel COM, Linux uses LibreOffice."""
    if platform.system() == "Windows":
        return _convert_xls_to_xlsx_windows(xls_path)
    return _convert_xls_to_xlsx_libreoffice(xls_path)


def _generate_dynamic_normalization_rules(approvers: List[Dict]) -> List[dict]:
    """Auto-generate normalization rules from approval role names.

    For each role like ``分管领导审核``, generate rules that swap the
    action suffix so that common variants (e.g. ``分管领导签字``) are
    normalised to the standard name.

    Manually configured rules (from ``payroll_sheet_config.json``) take
    precedence over these dynamic ones.
    """
    # Chinese approval action suffixes and their common synonyms
    _SUFFIX_SYNONYMS = {
        "审核": ["签字", "签章", "审批", "确认"],
        "签字": ["审核", "签章", "审批"],
        "审批": ["审核", "签字", "签章"],
        "签章": ["审核", "签字", "审批"],
    }
    _SUFFIX_LEN = 2  # all suffixes above are 2 characters

    rules: List[dict] = []
    seen: set = set()
    for approver in approvers:
        role = approver.get("role", "")
        if not role or len(role) < _SUFFIX_LEN or role in seen:
            continue
        seen.add(role)

        suffix = role[-_SUFFIX_LEN:]
        synonyms = _SUFFIX_SYNONYMS.get(suffix)
        if not synonyms:
            continue

        core = role[:-_SUFFIX_LEN]
        for variant_suffix in synonyms:
            variant = f"{core}{variant_suffix}"
            if variant != role:
                rules.append({"source": variant, "target": role})

    return rules


# ════════════════════════════════════════════════════════════════
# 公式求值器 —— 将公式原地替换为数值
# ════════════════════════════════════════════════════════════════

def _col_letter_to_num(col: str) -> int:
    """A→1, B→2, …, Z→26, AA→27, AB→28"""
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def _parse_cell_ref(ref: str):
    r"""Parse "A1" or "$A$1" into (row, col). Returns None on failure."""
    ref = ref.replace('$', '')
    m = re.match(r'^([A-Z]+)(\d+)$', ref)
    if not m:
        return None
    return int(m.group(2)), _col_letter_to_num(m.group(1))


def _resolve_cell_dependency(ws, row: int, col: int, visited: set):
    """
    Recursively resolve a cell to a numeric value.

    - Real number → return as-is
    - Formula → evaluate via ``_eval_simple_formula``
    - MergedCell / None / text → return *None* (cannot participate in arithmetic)
    """
    key = (row, col)
    if key in visited:
        return None
    visited.add(key)
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str) and val.startswith("="):
        return _eval_simple_formula(ws, val, visited)
    # Excel 中空单元格在算术/SUM 中等价于 0
    if val is None:
        return 0.0
    return None


def _eval_simple_formula(ws, formula: str, visited: set):
    """
    Evaluate a simple spreadsheet formula and return a numeric value.

    Supports: ``ROUND(expr, decimals)``, ``SUM(start:end)``,
    arithmetic (``+``, ``-``, ``*``, ``/``), parentheses, and cell references.

    Returns *None* when the formula cannot be evaluated (circular reference,
    unknown function, missing dependency, etc.).
    """
    text = formula.lstrip("=").strip()
    if not text:
        return None

    # ── ROUND(expr, decimals) ──
    m = re.match(r'^ROUND\((.+),\s*(\d+)\)$', text, re.IGNORECASE)
    if m:
        inner, decimals = m.groups()
        val = _eval_arithmetic_expr(ws, inner, set(visited))
        if val is not None:
            return round(val, int(decimals))

    # ── SUM(start:end) ──
    m = re.match(r'^SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', text, re.IGNORECASE)
    if m:
        col1, row1_s, col2, row2_s = m.groups()
        total = 0.0
        for r in range(int(row1_s), int(row2_s) + 1):
            for c in range(_col_letter_to_num(col1), _col_letter_to_num(col2) + 1):
                v = _resolve_cell_dependency(ws, r, c, set(visited))
                if v is None:
                    return None
                total += v
        return total

    # ── Plain arithmetic (e.g. J5*0.08, A5-B5-C5) ──
    return _eval_arithmetic_expr(ws, text, set(visited))


# ── tokenizer + recursive-descent arithmetic evaluator ──

_TOKEN_RE = re.compile(r'''
    [A-Z]+ \d+        # cell reference (A1, AB5)
    | \d+ \. \d+      # float
    | \d+             # integer
    | [+\-*/(),]      # operators & punctuation
    | \s+             # whitespace (ignored)
''', re.VERBOSE)


class _FormulaEvalError(Exception):
    """Raised when a formula cannot be evaluated."""


def _eval_arithmetic_expr(ws, expr: str, visited: set):
    """Entry point for arithmetic expression evaluation."""
    tokens = [t.strip() for t in _TOKEN_RE.findall(expr) if t.strip()]
    if not tokens:
        return None
    try:
        result, pos = _eval_add_sub(ws, tokens, 0, visited)
        if pos != len(tokens):
            return None
        return result
    except _FormulaEvalError:
        return None


def _eval_add_sub(ws, tokens, pos, visited):
    """+ / -  (lowest precedence)."""
    left, pos = _eval_mul_div(ws, tokens, pos, visited)
    while pos < len(tokens) and tokens[pos] in ('+', '-'):
        op = tokens[pos]
        right, pos = _eval_mul_div(ws, tokens, pos + 1, visited)
        left = left + right if op == '+' else left - right
    return left, pos


def _eval_mul_div(ws, tokens, pos, visited):
    """* /  (medium precedence)."""
    left, pos = _eval_unary(ws, tokens, pos, visited)
    while pos < len(tokens) and tokens[pos] in ('*', '/'):
        op = tokens[pos]
        right, pos = _eval_unary(ws, tokens, pos + 1, visited)
        if op == '*':
            left *= right
        else:
            if right == 0:
                raise _FormulaEvalError("division by zero")
            left /= right
    return left, pos


def _eval_unary(ws, tokens, pos, visited):
    """Unary minus / plus."""
    if pos < len(tokens) and tokens[pos] == '-':
        val, pos = _eval_primary(ws, tokens, pos + 1, visited)
        return -val, pos
    if pos < len(tokens) and tokens[pos] == '+':
        return _eval_primary(ws, tokens, pos + 1, visited)
    return _eval_primary(ws, tokens, pos, visited)


def _eval_primary(ws, tokens, pos, visited):
    """Number, cell reference, or parenthesised sub-expression."""
    if pos >= len(tokens):
        raise _FormulaEvalError("unexpected end")
    tok = tokens[pos]

    # Number
    if re.match(r'^\d+(\.\d+)?$', tok):
        return float(tok), pos + 1

    # Cell reference
    ref = _parse_cell_ref(tok)
    if ref:
        row, col = ref
        v = _resolve_cell_dependency(ws, row, col, set(visited))
        if v is None:
            raise _FormulaEvalError(f"cannot resolve {tok}")
        return float(v), pos + 1

    # Parenthesised expression
    if tok == '(':
        val, pos = _eval_add_sub(ws, tokens, pos + 1, visited)
        if pos >= len(tokens) or tokens[pos] != ')':
            raise _FormulaEvalError("missing ')'")
        return val, pos + 1

    raise _FormulaEvalError(f"unexpected token: {tok}")


def _replace_all_formulas_with_values(ws) -> None:
    """
    将工作表中所有可求值的公式替换为数值。

    扫描全部单元格，遇到公式就尝试求值：
    - 成功 → 原地替换为 float
    - 失败 → 保留原公式（记录 warn 日志）
    """
    replaced = 0
    failed = 0
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                val = _eval_simple_formula(ws, cell.value, set())
                if val is not None:
                    cell.value = val
                    replaced += 1
                else:
                    failed += 1
    if replaced:
        logger.info(f"[EVAL] Formula→value: {replaced} replaced, {failed} failed")
    elif failed:
        logger.info(f"[EVAL] All {failed} formulas have unresolved dependencies (kept as-is)")


def _estimate_text_render_width(text: str, font_size_pt: float) -> int:
    """Estimate the rendered pixel width of a piece of text in a given font size.

    Chinese / fullwidth characters are assumed to be roughly ``font_size`` px
    wide; ASCII / halfwidth characters are ``font_size * 0.6`` px.  The return
    value can be used as a ``OneCellAnchor.colOff`` EMU offset so an image
    starts immediately after the text.
    """
    px = 0.0
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff' or '\uff00' <= ch <= '\uffef':
            px += font_size_pt
        else:
            px += font_size_pt * 0.6
    return int(px * 914400 / 96)  # px → EMU at 96 DPI


def _insert_signature_to_excel_openpyxl(
    excel_path: Path,
    approvers: List[Dict],
    signatures_dir: Path,
    output_path: Path,
) -> Tuple[bool, List[str], Path]:
    inserted_roles = []
    try:
        logger.info(f"[SIGN] Loading workbook: {excel_path.name}")
        wb = load_workbook(str(excel_path))

        cfg = get_payroll_config()
        payroll_ws = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if is_payroll_sheet(ws, cfg):
                payroll_ws = ws
                logger.info(f"[SIGN] Found payroll sheet: {sheet_name}")
                break

        if payroll_ws is None:
            logger.warning(f"[SIGN] No payroll sheet found in {excel_path.name}")
            return False, [], output_path

        # 归一化文本单元格 —— 先应用动态规则（从审批角色自动派生），
        # 再叠加手动配置规则（后者优先级高）。
        dynamic_rules = _generate_dynamic_normalization_rules(approvers)
        config_rules = cfg.get("text_normalization", {}).get("rules", [])
        normalization_rules = dynamic_rules + config_rules
        if dynamic_rules:
            logger.info(f"[SIGN] Dynamic normalization rules: {dynamic_rules}")
        for row in range(1, payroll_ws.max_row + 1):
            for col in range(1, payroll_ws.max_column + 1):
                cell = payroll_ws.cell(row=row, column=col)
                if isinstance(cell.value, str):
                    normalized = _apply_normalization_rules(cell.value, normalization_rules)
                    if normalized != cell.value:
                        cell.value = normalized

        # 将全部公式原地求值为数值，后续列宽计算、列删除都基于真实值
        # 打印只输出值，公式无意义；且列宽按公式文本估算不准确会导致 ####
        _replace_all_formulas_with_values(payroll_ws)

        # 展开表头合并单元格 + 列删除 + 重建合并 ——
        # 仅对系统生成的工资表执行（行 2 有单位名称），手工表跳过列操作避免出错
        if _has_unit_name_in_row2(payroll_ws):

            # 必须在列删除前执行，否则合并范围地址会失准
            _flatten_header_merges(payroll_ws)
            _remove_force_delete_columns(payroll_ws, cfg)
            if cfg.get("remove_empty_columns", {}).get("enabled", True):
                _remove_empty_columns(payroll_ws, cfg)
            # 列删除后根据实际单元格值重新合并相邻相同表头
            _rebuild_header_merges(payroll_ws)
        else:
            logger.info(
                f"[SIGN] Manual sheet (no unit name in row 2), "
                f"column-deletion skipped for {excel_path.name}"
            )
        positions = find_all_signature_positions(payroll_ws, cfg)
        adjust_excel_for_print(payroll_ws, cfg)
        logger.info(f"[SIGN] Found positions: {positions}")
        if not positions:
            logger.warning(f"[SIGN] No signature positions found in {excel_path.name}")
            return False, [], output_path

        logger.info(f"[SIGN] Approvers: {[a['role'] for a in approvers]}")

        # ── Overwrite signature prompt cells with Feishu approval node names ──
        # Instead of relying on the original Excel hint text and normalization
        # rules, directly write the Feishu node name into each matched cell.
        # This eliminates dependency on original prompt text variants.
        for approver in approvers:
            role = approver.get("role")
            if role and role in positions:
                row, col = positions[role]
                old_val = payroll_ws.cell(row=row, column=col).value
                payroll_ws.cell(row=row, column=col).value = f"{role}："
                logger.info(f"[SIGN] Rewrote cell ({row},{col}) "
                            f"'{old_val}' → '{role}：'")

        for approver in approvers:
            role = approver.get("role")
            approver_name = approver.get("approver_name")
            if not role or not approver_name:
                logger.info(f"[SIGN] Skipping approver: missing role or approver_name")
                continue

            if role not in positions:
                logger.info(f"[SIGN] Role '{role}' not in positions")
                continue

            name_to_uid = _build_name_to_uid_mapping()
            uid = name_to_uid.get(approver_name, "")
            sig_path = get_signature_path(approver_name, signatures_dir, uid)
            if not sig_path:
                logger.warning(f"[SIGN] Signature image not found for {approver_name}")
                continue

            row, col = positions[role]
            _split_merged_for_text(payroll_ws, row, col)

            text_cell = payroll_ws.cell(row=row, column=col)
            text_cell.font = Font(size=10)

            # 根据提示词文本宽度计算图片偏移量，使签名紧跟在文字后方
            # 留两个字余量（约 20px），避免签名与提示词重叠
            text_val = str(text_cell.value) if text_cell.value else ""
            off_emu = _estimate_text_render_width(text_val, 10) + int(20 * 914400 / 96)

            img = XLImage(str(sig_path))
            img.width = 120
            img.height = 60
            anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=col - 1, row=row - 1,
                    colOff=off_emu, rowOff=0,
                ),
                ext=XDRPositiveSize2D(
                    pixels_to_EMU(img.width),
                    pixels_to_EMU(img.height),
                ),
            )
            payroll_ws.add_image(img, anchor)
            inserted_roles.append(role)
            logger.info(f"[SIGN] Inserted signature for {role} ({text_val}) offset={off_emu}")

        sig_rows = {r for r, _ in positions.values()}
        for r in sig_rows:
            for c in range(1, payroll_ws.max_column + 1):
                cell = payroll_ws.cell(row=r, column=c)
                if cell.value and "制表人" in str(cell.value):
                    old_size = cell.font.size or 11
                    if old_size > 10:
                        cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='right', vertical='center')

        _prevent_signature_page_split(payroll_ws, positions)

        actual_output = _build_output_path(excel_path, output_path, payroll_ws)

        sheets_to_remove = [sn for sn in wb.sheetnames if sn != payroll_ws.title]
        for sn in sheets_to_remove:
            del wb[sn]
            logger.info(f"[SIGN] Removed non-payroll sheet: {sn}")

        try:
            wb.save(str(actual_output))
        except PermissionError:
            fallback = actual_output.parent / f"{actual_output.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            logger.warning(f"[SIGN] 文件被占用，无法写入 {actual_output.name}，另存为 {fallback.name}")
            wb.save(str(fallback))
            actual_output = fallback
        logger.info(f"[SIGN] Saved to {actual_output.name}, inserted: {inserted_roles}")
        return len(inserted_roles) > 0, inserted_roles, actual_output
    except Exception as e:
        logger.error(f"[SIGN] Insertion failed: {e}")
        return False, [], output_path


def sanitize_dir_name(name: str) -> str:
    """Sanitize string for use as directory name."""
    invalid_chars = '\\/:*?"<>|'
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name.strip()


def get_signature_path(approver_name: str, signatures_dir: Path, user_id: str = "") -> Optional[Path]:
    """Look up signature PNG file for an approver, by user_id first then by name."""
    if user_id:
        sig_path = signatures_dir / f"{user_id}.png"
        if sig_path.exists():
            return sig_path
    if approver_name:
        sig_path = signatures_dir / f"{approver_name}.png"
        if sig_path.exists():
            return sig_path
        for f in signatures_dir.glob("*.png"):
            if approver_name in f.name or (user_id and user_id in f.name):
                return f
    return None


def _build_name_to_uid_mapping() -> dict:
    mapping_path = Path(__file__).parent / "user_mapping.json"
    try:
        with open(mapping_path, "r", encoding="utf-8") as f:
            forward = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    return {name: uid for uid, name in forward.items() if name}


def process_single_approval(
    instance_code: str,
    token: str,
    config: dict,
) -> dict:
    """Process a single approval: download, sign, print.

    Args:
        instance_code: Feishu approval instance code.
        token: Feishu tenant access token.
        config: Dict with keys:
            - save_dir: download directory (default: "./downloads")
            - signatures_dir: signature images directory (default: "./signatures")
            - role_mapping_path: optional path to role mapping json

    Returns:
        Result dict with keys: instance_code, success, message, downloaded,
        signed, signed_files, skipped, title.
    """
    result = {
        "instance_code": instance_code,
        "success": False,
        "message": "",
        "downloaded": [],
        "signed": [],
        "signed_files": [],
        "skipped": False,
        "title": "",
        "approval_code": "",
        "cashier_task": None,
    }

    try:
        logger.info(f"[BATCH] Getting details for {instance_code}...")
        detail = get_instance_detail(token, instance_code)
        logger.info(f"[BATCH] Got details from API")
    except Exception as e:
        logger.error(f"[BATCH] Failed to get details: {e}")
        result["message"] = f"获取详情失败: {e}"
        return result

    result["title"] = detail.get("approval_name", instance_code[:20])
    result["approval_code"] = detail.get("approval_code", "")

    # 提取待审批的出纳办理任务信息
    for task in detail.get("task_list", []) or []:
        if task.get("node_name") == "出纳办理" and task.get("status") == "PENDING":
            result["cashier_task"] = {
                "task_id": task.get("id"),
                "open_id": task.get("open_id"),
            }
            break

    role_mapping_path = config.get("role_mapping_path")
    if role_mapping_path:
        role_mapping_path = Path(role_mapping_path)
    approvers = get_approvers_with_roles(detail, role_mapping_path)
    approvers = [a for a in approvers if a.get("status") == "APPROVED" and a.get("role")]

    form_widgets = parse_form(detail)
    attachments = extract_attachments(form_widgets)
    if not attachments:
        result["message"] = "无附件"
        return result

    save_dir = Path(config.get("save_dir", "./downloads"))
    serial = detail.get("serial_number") or instance_code
    name_to_uid = _build_name_to_uid_mapping()
    uid = detail.get("user_id", "")
    submitter_name = uid
    for name, mid in name_to_uid.items():
        if mid == uid:
            submitter_name = name
            break
    form_title = ""
    for w in form_widgets:
        if w.get("name") == "标题":
            form_title = w.get("value", "")
            break
    dir_name = f"{serial}_{submitter_name}_{form_title}" if form_title else f"{serial}_{submitter_name}"
    dir_name = sanitize_dir_name(dir_name) or sanitize_dir_name(instance_code)
    instance_dir = save_dir / dir_name
    instance_dir.mkdir(parents=True, exist_ok=True)

    signatures_dir = Path(config.get("signatures_dir", "./signatures"))

    for att in attachments:
        field_name = att.get("field_name", "附件")
        values = att.get("value", [])
        if not values:
            continue

        if "汇总表" in field_name:
            logger.info(f"[BATCH] Skipping summary table: {field_name}")
            continue

        for file_token in values:
            try:
                logger.info(f"[BATCH] Processing attachment: {field_name}")
                downloaded_path = download_file(token, file_token, str(instance_dir))
                file_path = Path(downloaded_path)
                result["downloaded"].append(file_path.name)

                if file_path.suffix.lower() in (".xlsx", ".xls"):
                    if approvers:
                        sign_path = file_path
                        if file_path.suffix.lower() == ".xls":
                            converted = _convert_xls_to_xlsx(file_path)
                            if converted:
                                sign_path = converted
                            else:
                                logger.warning(f"[BATCH] Cannot convert .xls, skipping signature: {file_path.name}")
                                continue

                        logger.info(f"[BATCH] Inserting signatures into {sign_path.name}...")
                        signed_name = f"signed_{sign_path.stem}.xlsx"
                        signed_path = instance_dir / signed_name
                        success, inserted, actual_signed_path = _insert_signature_to_excel_openpyxl(
                            sign_path, approvers, signatures_dir, signed_path
                        )
                        if success:
                            logger.info(f"[BATCH] Signature insertion success: {inserted}")
                            result["signed"].extend(inserted)
                            result["signed_files"].append(str(actual_signed_path))
                        else:
                            logger.warning(f"[BATCH] Signature insertion failed for {file_path.name}")
                    else:
                        logger.info(f"[BATCH] No approved approvers with roles, skipping signature: {file_path.name}")
                else:
                    logger.info(f"[BATCH] Non-Excel file, skipping signature: {file_path.name}")

            except Exception as e:
                logger.error(f"[BATCH] Error processing {field_name}: {e}")
                continue

    result["success"] = True
    result["message"] = (
        f"下载 {len(result['downloaded'])} 个, "
        f"签名 {len(result['signed'])} 处"
    )
    return result
