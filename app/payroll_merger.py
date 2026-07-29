"""
Payroll merger and WPS print functions.

Extracted and adapted from seqprint's batchprint_gui.py.
Provides payroll file merging by big-org grouping, WPS COM printing,
and mapping-rule based unit-to-org resolution.
"""

import logging
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Callable, List, Optional, Tuple

import openpyxl

logger = logging.getLogger(__name__)

# ── Helpers for column auto-sizing ──
def _is_numeric_com(v) -> bool:
    """Check if a COM cell value represents a number (int/float or numeric string)."""
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v)
            return True
        except (ValueError, TypeError):
            pass
    return False

def _com_estimate_width(val) -> float:
    """Estimate WPS column-width units needed to display ``val`` in 宋体 9pt."""
    s = str(val) if val is not None else ""
    w = 0.0
    for ch in s:
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f':
            w += 2.0
        else:
            w += 1.0
    return w

# ── Global caches for mapping rules ──
_MAPPING_RULES_CACHE: Optional[List[Tuple[str, str, str, bool]]] = None
_MAPPING_EXCEL: Optional[Path] = None


# ════════════════════════════════════════════════════════════════
#  Mapping-rule helpers
# ════════════════════════════════════════════════════════════════

def _get_mapping_excel_path() -> Path:
    """Return the absolute path to the mapping rules Excel file."""
    global _MAPPING_EXCEL
    if _MAPPING_EXCEL is None:
        _MAPPING_EXCEL = Path(__file__).parent / "template" / "映射规则.xlsx"
    return _MAPPING_EXCEL


def _load_mapping_rules() -> List[Tuple[str, str, str, bool]]:
    """
    Read mapping rules from *映射规则.xlsx*.

    Returns a list of tuples ``(match_type, pattern, big_org, is_excluded)``:

    =========== =============================================================
    Field       Description
    =========== =============================================================
    match_type  ``"prefix"`` / ``"contains"`` / ``"exact"``
    pattern     The string to match against the unit name
    big_org     The big-organisation name this unit belongs to (empty str if
                the row only marks the unit as excluded / unmapped)
    is_excluded ``True`` when the unit should be excluded from grouping
    =========== =============================================================

    Results are cached after the first call.
    """
    global _MAPPING_RULES_CACHE
    if _MAPPING_RULES_CACHE is not None:
        return _MAPPING_RULES_CACHE

    excel_path = _get_mapping_excel_path()
    if not excel_path.exists():
        _MAPPING_RULES_CACHE = []
        return _MAPPING_RULES_CACHE

    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    except Exception:
        logger.warning("无法打开映射规则文件: %s", excel_path)
        _MAPPING_RULES_CACHE = []
        return _MAPPING_RULES_CACHE

    if "映射规则" not in wb.sheetnames:
        wb.close()
        _MAPPING_RULES_CACHE = []
        return _MAPPING_RULES_CACHE

    ws = wb["映射规则"]
    rules: List[Tuple[str, str, str, bool]] = []
    for r in range(2, ws.max_row + 1):
        match_type = ws.cell(r, 2).value   # B column
        pattern = ws.cell(r, 3).value       # C column
        big_org = ws.cell(r, 4).value       # D column
        excluded = ws.cell(r, 5).value      # E column
        if not pattern or not match_type:
            continue
        rules.append((
            str(match_type).strip(),
            str(pattern).strip(),
            str(big_org).strip() if big_org else "",
            str(excluded).strip() == "Y",
        ))

    wb.close()
    _MAPPING_RULES_CACHE = rules
    logger.info("加载映射规则: %d 条", len(rules))
    return rules


def _normalize_unit_name(name: str) -> str:
    """
    Strip Windows file-copy suffixes from a unit name.

    Removes in order: ``（1）``, ``(2)``, ``- 副本``, ``- Copy``, ``- 复件``,
    ``_副本``, etc.
    """
    s = str(name).strip()
    patterns = [
        r'\s*[（(]\d+[）)]\s*$',            # （1） (2)
        r'\s*[-–—]\s*副本\s*$',              # - 副本  — 副本
        r'\s*[-–—]\s*[Cc]opy\s*$',           # - Copy  - copy
        r'\s*[-–—]\s*复件\s*$',              # - 复件
        r'\s*_副本\s*$',                     # _副本
    ]
    for p in patterns:
        s = re.sub(p, '', s)
    return s.strip()


def is_excluded(unit_name: str) -> bool:
    """
    Check whether *unit_name* is in the exclusion list.

    A unit is considered excluded when any mapping rule with
    ``is_excluded=True`` matches it.
    """
    rules = _load_mapping_rules()
    for match_type, pattern, _big_org, excluded in rules:
        if not excluded:
            continue
        if match_type == "prefix" and unit_name.startswith(pattern):
            return True
        if match_type == "exact" and unit_name == pattern:
            return True
        if match_type == "contains" and pattern in unit_name:
            return True
    return False


def get_big_org(unit_name: str) -> Tuple[str, str]:
    """
    Resolve a unit name to its big organisation via mapping rules.

    Args:
        unit_name: The raw unit name (e.g. from filename or sheet content).

    Returns:
        A tuple ``(big_org, matched_unit_name)``.  When a rule matches,
        *big_org* is the mapped org name (or the original *unit_name* if
        the rule has no big_org); *matched_unit_name* is the original
        *unit_name*.  Falls back to ``(unit_name, unit_name)`` when no
        rule matches.
    """
    rules = _load_mapping_rules()
    for match_type, pattern, big_org, excluded in rules:
        matched = False
        if match_type == "prefix" and unit_name.startswith(pattern):
            matched = True
        elif match_type == "exact" and unit_name == pattern:
            matched = True
        elif match_type == "contains" and pattern in unit_name:
            matched = True
        if not matched:
            continue
        if big_org:
            return big_org, unit_name
        return unit_name, unit_name
    # No match — fallback to the original name
    return unit_name, unit_name


# ════════════════════════════════════════════════════════════════
#  Filename / sheet helpers
# ════════════════════════════════════════════════════════════════

def _extract_unit_from_signed(fname: str) -> str:
    """
    Extract the unit name from a ``signed_``-prefixed filename.

    Strips the ``signed_`` prefix and removes the year-month suffix
    (e.g. ``202606``, ``2026年06月``).
    """
    base = fname[len("signed_"):] if fname.startswith("signed_") else fname
    # Match year-month patterns: 202606, 2026年06月, 2026年6月
    m = re.search(r'\d{4}[年]?\d{1,2}[月]?', base)
    if m:
        return base[:m.start()].strip().rstrip("-—_")
    # Fallback: strip .xlsx extension
    return base.rsplit(".", 1)[0].strip()


def _extract_yearmon_from_signed(fname: str) -> str:
    """
    Extract a ``YYYYMM`` string from a ``signed_``-prefixed filename.

    Returns an empty string when no year-month pattern is found.
    """
    m = re.search(r'(\d{4})[年]?(\d{1,2})[月]?', fname)
    if m:
        return m.group(1) + m.group(2).zfill(2)
    return ""


def _format_yearmons(yearmon_set: set) -> str:
    """
    Format a set of year-month strings for display.

    Consecutive months are joined with ``-``, non-consecutive with ``、``.

    Example: ``{"202601", "202602", "202604"}`` → ``"202601-202602、202604"``
    """
    sorted_ym = sorted(set(ym for ym in yearmon_set if ym))
    if not sorted_ym:
        return ""
    groups = []
    cur = [sorted_ym[0]]
    for ym in sorted_ym[1:]:
        prev = int(cur[-1])
        cur_i = int(ym)
        if (cur_i == prev + 1
                or (cur_i % 100 == 1 and prev % 100 == 12
                    and cur_i // 100 == prev // 100 + 1)):
            cur.append(ym)
        else:
            groups.append(cur)
            cur = [ym]
    groups.append(cur)
    parts = []
    for g in groups:
        if len(g) >= 2:
            parts.append(f"{g[0]}-{g[-1]}")
        else:
            parts.append(g[0])
    return "、".join(parts)


# ════════════════════════════════════════════════════════════════
#  WPS COM helpers
# ════════════════════════════════════════════════════════════════

# ── WPS COM error tracking ────────────────────────────────────────
_wps_last_error: str = ""

def _set_wps_error(msg: str) -> None:
    global _wps_last_error
    _wps_last_error = msg

def get_last_wps_error() -> str:
    """Return the last WPS COM check error message (empty if OK)."""
    return _wps_last_error


def check_wps_available() -> bool:
    """
    Check whether the WPS Office COM component (``KET.Application``)
    is available on the current machine.

    Returns ``True`` when WPS is reachable, ``False`` otherwise
    (including on non-Windows platforms).
    """
    try:
        import pythoncom
        pythoncom.CoInitialize()
        import win32com.client  # noqa: F811

        app = win32com.client.DispatchEx("KET.Application")
        app.Quit()
        return True
    except ImportError:
        _set_wps_error("pywin32 未安装，不在 Windows 环境")
        return False
    except Exception as e:
        msg = f"{type(e).__name__}: {e}"
        logger.warning("WPS COM check failed: %s", msg)
        _set_wps_error(msg)
        return False


def print_file(
    filepath: str,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
    repeat_header: bool = True,
) -> bool:
    """
    Print a single Excel file via WPS COM.

    Opens the file, sets page setup (A4 landscape, fit-to-page wide,
    optional repeating title rows), and sends it to the printer.
    Retries up to 3 times on failure.

    Args:
        filepath:        Path to the Excel file.
        progress_callback: Optional ``(current, total, message)`` callback.
        repeat_header:   When ``True``, ``PrintTitleRows = $1:$5`` is set
                         so the header repeats on every page.

    Returns:
        ``True`` on successful print, ``False`` otherwise.
    """
    import time

    try:
        import win32com.client  # noqa: F811
        import pythoncom
        pythoncom.CoInitialize()
    except ImportError:
        if progress_callback:
            progress_callback(0, 1, "win32com 不可用（非 Windows 环境）")
        return False

    max_retries = 3
    for attempt in range(1, max_retries + 1):
        app = None
        try:
            app = win32com.client.DispatchEx("KET.Application")
            app.Visible = False
            app.DisplayAlerts = False

            wb = app.Workbooks.Open(filepath)
            ws = wb.ActiveSheet

            # Page setup: A4 landscape, fit all columns to 1 page wide
            ws.PageSetup.Orientation = 2          # xlLandscape
            ws.PageSetup.PaperSize = 9            # xlPaperA4
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = 0
            if repeat_header:
                ws.PageSetup.PrintTitleRows = "$1:$5"
            # else: leave PrintTitleRows as-is — the file was saved
            # without PrintTitleRows during merge, so the blank
            # setting is already correct.  WPS COM may ignore an
            # explicit "" assignment on a property that was already
            # empty at file-open time, causing stale cached headers
            # from the template to reappear.

            # Fix picture row height so images are not compressed
            try:
                for _shape in ws.Shapes:
                    if _shape.Type == 13:  # msoPicture
                        _row = _shape.TopLeftCell.Row
                        ws.Rows(_row).RowHeight = 50
            except Exception:
                pass

            ws.PrintOut()
            wb.Close(SaveChanges=False)
            return True

        except Exception as e:
            import traceback as _tb

            err_detail = f"{type(e).__name__}: {e}"
            if progress_callback:
                progress_callback(
                    0, 1,
                    f"打印失败 (第{attempt}次): {err_detail}"
                )
                progress_callback(
                    0, 1,
                    f"  traceback: {_tb.format_exc()[:500]}"
                )
            if attempt < max_retries:
                time.sleep(5)
            else:
                return False
        finally:
            if app is not None:
                try:
                    app.Quit()
                except Exception:
                    pass
            time.sleep(2.5)


def batch_print(
    file_list: List[str],
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> Tuple[int, int, List[str]]:
    """
    Batch-print a list of Excel files via WPS COM.

    Args:
        file_list:        List of file paths to print.
        progress_callback: Optional ``(current, total, message)`` callback.

    Returns:
        ``(success_count, fail_count, fail_list)`` — *fail_list* contains
        the file paths that failed to print.
    """
    if not check_wps_available():
        return 0, 0, ["WPS不可用"]

    total = len(file_list)
    success_count = 0
    fail_count = 0
    fail_list: List[str] = []

    for i, filepath in enumerate(file_list):
        if progress_callback:
            progress_callback(i + 1, total, f"正在打印: {filepath}")

        # Merged "合集" files already contain embedded headers in each
        # source sheet — do NOT repeat virtual headers on every page.
        is_merged = "_工资表合集" in Path(filepath).name
        ok = print_file(filepath, progress_callback, repeat_header=not is_merged)
        if ok:
            success_count += 1
        else:
            fail_count += 1
            fail_list.append(filepath)

    return success_count, fail_count, fail_list


# ════════════════════════════════════════════════════════════════
#  Merge pipeline
# ════════════════════════════════════════════════════════════════

def _extract_unit_from_sheet(fpath: str) -> str:
    """
    Extract the unit name from a payroll workbook's row 2.

    This is more reliable than filename-based extraction because the
    sheet content always carries the official unit name.

    Returns an empty string when the unit name cannot be determined.
    """
    try:
        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        for c in range(1, 4):
            v = ws.cell(row=2, column=c).value
            if v and "名称" in str(v):
                # "名称：吉林大学XX学院" → content after colon
                m = re.search(r'[：:]\s*(.*)', str(v))
                if m and m.group(1).strip():
                    return m.group(1).strip()
                # "名称" in C1, "：吉林大学XX学院" in C2
                nv = ws.cell(row=2, column=c + 1).value
                if nv and nv.strip().lstrip("：: "):
                    return str(nv).strip().lstrip("：: ")
        wb.close()
    except Exception:
        logger.warning("无法从文件内容提取单位名: %s", fpath)
    return ""


def merge_payrolls_simple(
    payroll_dir: str,
    output_dir: str,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
    left_margin_cm: float = 1.5,
    right_margin_cm: float = 0.5,
) -> Tuple[List[str], List[str], dict]:
    """
    Merge payroll worksheets by big organisation.

    Scans all ``signed_*.xlsx`` files under *payroll_dir* (recursive),
    extracts the unit name from each sheet, resolves it to a big org
    via mapping rules, groups files by org, then produces one merged
    Excel file per group.

    Each output file contains:
      1. A **virtual summary table** at the top with the org name,
         year-month range, unified column headers, per-unit totals,
         and a grand-total row.
      2. The **original worksheets** stacked below (including signature
         images), pasted via WPS COM cross-workbook copy to preserve
         formatting.

    Args:
        payroll_dir:       Directory to scan for ``signed_*.xlsx`` files.
        output_dir:        Directory for the generated merged files.
        progress_callback: Optional ``(current, total, message)`` callback.
        left_margin_cm:    Left page margin in centimetres (default 1.5).
        right_margin_cm:   Right page margin in centimetres (default 0.5).

    Returns:
        ``(output_files, warnings, stats)``:
        - *output_files* — paths of the generated files.
        - *warnings* — list of warning messages.
        - *stats* — dict with key ``"total_groups"``.
    """
    import win32com.client  # noqa: F811
    import pythoncom
    pythoncom.CoInitialize()

    warnings_list: List[str] = []

    # ── 1. Scan directory ──
    signed_files: List[str] = []
    for root, _dirs, files in os_walk(str(payroll_dir)):
        for f in files:
            if f.startswith("signed_") and f.endswith(".xlsx") and not f.startswith("~$"):
                signed_files.append(Path(root, f).as_posix())
    signed_files.sort()

    if not signed_files:
        return [], ["未找到 signed_*.xlsx 文件"], {}

    # ── 2. Extract info ──
    file_infos = []
    for fpath in signed_files:
        fname = Path(fpath).name
        unit_name = _extract_unit_from_sheet(fpath) or _extract_unit_from_signed(fname)
        yearmon = _extract_yearmon_from_signed(fname)
        big_org, _ = get_big_org(unit_name)
        excluded = is_excluded(unit_name)
        file_infos.append({
            "path": fpath,
            "fname": fname,
            "unit_name": unit_name,
            "big_org": big_org,
            "excluded": excluded,
            "yearmon": yearmon,
        })

    # ── 3. Group by big org ──
    groups: dict = defaultdict(list)
    group_display: dict = {}  # internal_key → display name for title/filename
    excl_counter = 0
    for info in file_infos:
        if info["excluded"]:
            excl_counter += 1
            # Each excluded file gets its own group (no sibling merging)
            key = f"__excl_{excl_counter}"
            group_display[key] = info["unit_name"]
        else:
            key = info["big_org"]
        if not key or not key.strip():
            key = Path(info["fname"]).stem  # fallback to filename stem
        groups[key].append(info)

    def _disp(gk: str) -> str:
        """User-facing group name (hide internal keys)."""
        return group_display.get(gk, gk)

    if progress_callback:
        progress_callback(
            0, len(groups),
            f"共 {len(signed_files)} 个工资表，分为 {len(groups)} 组",
        )

    output_files: List[str] = []
    _name_counter: dict = {}  # dedup suffix for same-named excluded files

    # ── 4. Start WPS ──
    app = None
    try:
        app = win32com.client.DispatchEx("KET.Application")
        app.Visible = False
        app.DisplayAlerts = False

        for group_idx, (group_key, items) in enumerate(groups.items(), 1):
            if progress_callback:
                progress_callback(
                    group_idx, len(groups),
                    f"合并组: {_disp(group_key)}（{len(items)} 个工资表）",
                )

            all_yearmons: set = set(info["yearmon"] for info in items)
            ym_display = _format_yearmons(all_yearmons)

            # Count columns for each source file
            max_cols = 0
            for info in items:
                src_wb_check = openpyxl.load_workbook(info["path"])
                src_ws_check = src_wb_check.active
                info["ncols"] = src_ws_check.max_column or 1
                info["nrows"] = src_ws_check.max_row or 1
                max_cols = max(max_cols, info["ncols"])
                src_wb_check.close()

            # ── Create target workbook ──
            tgt_wb = app.Workbooks.Add()
            tgt_ws = tgt_wb.ActiveSheet
            safe_name = re.sub(r'[\\/?*\[\]:]', '', _disp(group_key))[:31] or "未命名"
            tgt_ws.Name = safe_name

            # ── Read totals & column fingerprints per source file ──
            file_totals = []
            file_fingerprints = []
            for info in items:
                row_vals = {}
                fp_dict = {}
                try:
                    src_rb = openpyxl.load_workbook(info["path"], data_only=True)
                    src_ws = src_rb.active
                    src_nrows = src_ws.max_row or 1
                    src_ncols = src_ws.max_column or 1

                    # Build merge-lookup table for rows 3-5
                    merge_lookup = {}
                    for mc in src_ws.merged_cells.ranges:
                        mr1, mr2 = mc.min_row, mc.max_row
                        mc1, mc2 = mc.min_col, mc.max_col
                        if mr2 >= 3 and mr1 <= 5:
                            tl = src_ws.cell(row=mc.min_row, column=mc.min_col).value
                            if tl is not None:
                                tl_s = str(tl).strip()
                                for rr in range(max(mr1, 3), min(mr2, 5) + 1):
                                    for cc in range(mc1, mc2 + 1):
                                        merge_lookup[(rr, cc)] = tl_s

                    def _cv(row: int, col: int) -> str:
                        if (row, col) in merge_lookup:
                            return merge_lookup[(row, col)]
                        v = src_ws.cell(row=row, column=col).value
                        return str(v).strip() if v is not None else ""

                    for c in range(1, src_ncols + 1):
                        fp_dict[c] = (_cv(3, c), _cv(4, c), _cv(5, c))

                    # Find total row
                    for check_r in range(src_nrows, 0, -1):
                        v = src_ws.cell(row=check_r, column=1).value
                        if v is not None and str(v).strip() == "合计":
                            for c in range(1, src_ncols + 1):
                                cv = src_ws.cell(row=check_r, column=c).value
                                row_vals[c] = cv
                            break
                    src_rb.close()
                except Exception:
                    pass
                file_totals.append(row_vals)
                file_fingerprints.append(fp_dict)

            # ── Compute first data column from 合计 row ──
            data_start_col = max_cols + 1
            for row_vals in file_totals:
                if not row_vals:
                    continue
                for c in range(2, max_cols + 1):
                    if c in row_vals and row_vals[c] is not None:
                        data_start_col = min(data_start_col, c)
                        break
            if data_start_col > max_cols:
                data_start_col = 5

            # ── Build canonical column-fingerprint sequence ──
            variant_counts = []
            for fpf in file_fingerprints:
                cnt = sum(1 for c, fp in fpf.items() if c >= 3 and fp != ("", "", ""))
                variant_counts.append(cnt)
            count_freq = {}
            for c_val in variant_counts:
                count_freq[c_val] = count_freq.get(c_val, 0) + 1
            most_common_cnt = max(count_freq, key=count_freq.get)
            ref_idx = next(idx for idx, c_val in enumerate(variant_counts) if c_val == most_common_cnt)
            ref_fp = file_fingerprints[ref_idx]

            canonical_fps = []
            seen_fp = set()
            for c in range(data_start_col, max_cols + 1):
                fp = ref_fp.get(c, ("", "", ""))
                if fp != ("", "", "") and fp not in seen_fp:
                    canonical_fps.append(fp)
                    seen_fp.add(fp)
            for fpf in file_fingerprints:
                for c, fp in fpf.items():
                    if c >= data_start_col and fp != ("", "", "") and fp not in seen_fp:
                        seen_fp.add(fp)
                        insert_idx = len(canonical_fps)
                        same_r3v_last = -1
                        for idx, existing_fp in enumerate(canonical_fps):
                            if existing_fp[0] == fp[0]:
                                same_r3v_last = idx
                                if "合计" in existing_fp[1]:
                                    insert_idx = idx
                                    break
                        if insert_idx == len(canonical_fps) and same_r3v_last >= 0:
                            insert_idx = same_r3v_last + 1
                        if insert_idx == len(canonical_fps):
                            for lc in range(c - 1, data_start_col - 1, -1):
                                lfp = fpf.get(lc)
                                if lfp and lfp in canonical_fps:
                                    insert_idx = canonical_fps.index(lfp) + 1
                                    break
                            if insert_idx == len(canonical_fps):
                                for rc in range(c + 1, max_cols + 1):
                                    rfp = fpf.get(rc)
                                    if rfp and rfp in canonical_fps:
                                        insert_idx = canonical_fps.index(rfp)
                                        break
                        canonical_fps.insert(insert_idx, fp)

            # ── Build column maps per file ──
            file_col_maps = []
            for fpf in file_fingerprints:
                rev = {fp: src_c for src_c, fp in fpf.items()}
                col_map = {}
                for vi, fp in enumerate(canonical_fps):
                    if fp in rev:
                        col_map[vi] = rev[fp]
                file_col_maps.append(col_map)

            # ── Filter out columns with all-zero totals ──
            active_fps = []
            for vi, fp in enumerate(canonical_fps):
                has_data = False
                for ft, cm in zip(file_totals, file_col_maps):
                    src_c = cm.get(vi)
                    if src_c is None:
                        continue
                    v = ft.get(src_c)
                    if v is None:
                        continue
                    try:
                        if float(v) != 0:
                            has_data = True
                            break
                    except (ValueError, TypeError):
                        if str(v).strip():
                            has_data = True
                            break
                if has_data:
                    active_fps.append(fp)
            if not active_fps:
                active_fps = canonical_fps[:1]
            canonical_fps = active_fps

            # Rebuild column maps after filtering
            file_col_maps = []
            for fpf in file_fingerprints:
                rev = {fp: sc for sc, fp in fpf.items()}
                cm = {}
                for vi, fp in enumerate(canonical_fps):
                    if fp in rev:
                        cm[vi] = rev[fp]
                file_col_maps.append(cm)

            virtual_cols = 2 + len(canonical_fps)

            # ── Read reference file headers (for display) ──
            ref_hdr = [[""] * (max_cols + 1) for _ in range(3)]
            try:
                from openpyxl.utils import range_boundaries
                ref_wb = openpyxl.load_workbook(items[ref_idx]["path"])
                ref_ws = ref_wb.active
                for hi, hr in enumerate([3, 4, 5]):
                    for c in range(1, max_cols + 1):
                        v = ref_ws.cell(row=hr, column=c).value
                        if v is not None:
                            ref_hdr[hi][c] = str(v).strip()
                for mr in ref_ws.merged_cells.ranges:
                    mc, mr0, Mc, Mr = range_boundaries(str(mr))
                    if Mr < 3 or mr0 > 5:
                        continue
                    for hi, hr in enumerate(range(max(mr0, 3), min(Mr, 5) + 1)):
                        tl = ref_hdr[hr - 3][mc]
                        if not tl:
                            continue
                        for c in range(mc, Mc + 1):
                            ref_hdr[hr - 3][c] = tl
                ref_wb.close()
            except Exception:
                pass

            display_hdr = {}
            for c in range(data_start_col, max_cols + 1):
                fp = ref_fp.get(c)
                if fp and fp != ("", "", ""):
                    display_hdr[fp] = (
                        ref_hdr[0][c],
                        ref_hdr[1][c],
                        ref_hdr[2][c],
                    )

            # ── Write header rows ──
            r = 1
            # Row 1: title
            title = f"{_disp(group_key)} 工资表合集"
            if ym_display:
                title += f"（{ym_display}）"
            tgt_ws.Cells(r, 1).Value = title
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, virtual_cols)
            ).Merge()
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, virtual_cols)
            ).Font.Bold = True
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, virtual_cols)
            ).Font.Size = 16
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, virtual_cols)
            ).Font.Name = "宋体"
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, virtual_cols)
            ).HorizontalAlignment = -4108
            r += 1

            # Row 2: unit name + date
            right_start = virtual_cols - 3 if virtual_cols > 6 else 4
            tgt_ws.Cells(r, 1).Value = f"单位名称：{_disp(group_key)}"
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, right_start - 1)
            ).Merge()
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, right_start - 1)
            ).HorizontalAlignment = 1
            tgt_ws.Cells(r, right_start).Value = (
                f"统计时间：{datetime.now().strftime('%Y年%m月%d日')}"
            )
            tgt_ws.Range(
                tgt_ws.Cells(r, right_start), tgt_ws.Cells(r, virtual_cols)
            ).Merge()
            tgt_ws.Range(
                tgt_ws.Cells(r, right_start), tgt_ws.Cells(r, virtual_cols)
            ).HorizontalAlignment = -4152
            r += 1

            # Rows 3-5: composite column headers
            hdr_start_row = r
            for hi in range(3):
                tgt_ws.Cells(r, 1).Value = "序号" if hi == 0 else ""
                if hi == 0:
                    tgt_ws.Cells(r, 1).Font.Bold = True
                tgt_ws.Cells(r, 2).Value = "结算单元名称" if hi == 0 else ""
                if hi == 0:
                    tgt_ws.Cells(r, 2).Font.Bold = True
                for vi, fp_val in enumerate(canonical_fps, 3):
                    dv = display_hdr.get(fp_val, fp_val)[hi]
                    tgt_ws.Cells(r, vi).Value = dv
                    if hi == 0 and dv:
                        tgt_ws.Cells(r, vi).Font.Bold = True
                r += 1
            hdr_end_row = r - 1

            for hr in range(hdr_start_row, hdr_end_row + 1):
                tgt_ws.Range(
                    tgt_ws.Cells(hr, 1), tgt_ws.Cells(hr, virtual_cols)
                ).HorizontalAlignment = -4108

            # ── Merge header cells ──
            hmerged = set()
            # Row 4 horizontal merge (within row-3 groups)
            vi = 3
            while vi <= virtual_cols:
                r3v = tgt_ws.Cells(hdr_start_row, vi).Value
                vj = vi + 1
                while vj <= virtual_cols and tgt_ws.Cells(hdr_start_row, vj).Value == r3v:
                    vj += 1
                vk = vi
                while vk < vj:
                    r4v = tgt_ws.Cells(hdr_start_row + 1, vk).Value
                    if not r4v:
                        vk += 1
                        continue
                    vl = vk + 1
                    while vl < vj and tgt_ws.Cells(hdr_start_row + 1, vl).Value == r4v:
                        vl += 1
                    if vl - 1 > vk:
                        try:
                            tgt_ws.Range(
                                tgt_ws.Cells(hdr_start_row + 1, vk),
                                tgt_ws.Cells(hdr_start_row + 1, vl - 1),
                            ).Merge()
                            for vc in range(vk, vl):
                                hmerged.add((hdr_start_row + 1, vc))
                        except Exception:
                            pass
                    vk = vl
                vi = vj

            # Vertical merges
            for vi in range(3, virtual_cols + 1):
                r3v = tgt_ws.Cells(hdr_start_row, vi).Value
                r4v = tgt_ws.Cells(hdr_start_row + 1, vi).Value
                r5v = tgt_ws.Cells(hdr_start_row + 2, vi).Value
                if r4v and (not r5v or r4v == r5v):
                    if r3v and r3v == r4v:
                        if not any((rr, vi) in hmerged for rr in range(hdr_start_row, hdr_end_row + 1)):
                            try:
                                tgt_ws.Range(
                                    tgt_ws.Cells(hdr_start_row, vi),
                                    tgt_ws.Cells(hdr_end_row, vi),
                                ).Merge()
                                for rr in range(hdr_start_row, hdr_end_row + 1):
                                    hmerged.add((rr, vi))
                            except Exception:
                                pass
                    else:
                        if not any((rr, vi) in hmerged for rr in range(hdr_start_row + 1, hdr_end_row + 1)):
                            try:
                                tgt_ws.Range(
                                    tgt_ws.Cells(hdr_start_row + 1, vi),
                                    tgt_ws.Cells(hdr_end_row, vi),
                                ).Merge()
                                for rr in range(hdr_start_row + 1, hdr_end_row + 1):
                                    hmerged.add((rr, vi))
                            except Exception:
                                pass
            # Catch-all: rows 4+5 empty → merge 3 rows
            for vi in range(1, virtual_cols + 1):
                if any((rr, vi) in hmerged for rr in range(hdr_start_row, hdr_end_row + 1)):
                    continue
                r4v = tgt_ws.Cells(hdr_start_row + 1, vi).Value
                r5v = tgt_ws.Cells(hdr_start_row + 2, vi).Value
                if not r4v and not r5v:
                    try:
                        tgt_ws.Range(
                            tgt_ws.Cells(hdr_start_row, vi),
                            tgt_ws.Cells(hdr_end_row, vi),
                        ).Merge()
                    except Exception:
                        pass

            # Row 3 horizontal merge (last)
            vi = 3
            while vi <= virtual_cols:
                r3v = tgt_ws.Cells(hdr_start_row, vi).Value
                if not r3v:
                    vi += 1
                    continue
                vj = vi + 1
                while vj <= virtual_cols and tgt_ws.Cells(hdr_start_row, vj).Value == r3v:
                    vj += 1
                if vj - 1 > vi:
                    try:
                        tgt_ws.Range(
                            tgt_ws.Cells(hdr_start_row, vi),
                            tgt_ws.Cells(hdr_start_row, vj - 1),
                        ).Merge()
                    except Exception:
                        pass
                vi = vj

            data_start_row = r

            # ── Write data rows (per-file totals) ──
            total_accum = {}
            for idx, (info, ft) in enumerate(zip(items, file_totals), 1):
                tgt_ws.Cells(r, 1).Value = idx
                tgt_ws.Cells(r, 2).Value = info["unit_name"]
                col_map = file_col_maps[idx - 1]
                for vi, fp_val in enumerate(canonical_fps):
                    src_c = col_map.get(vi)
                    if src_c is not None and src_c in ft:
                        cv = ft[src_c]
                        if cv is not None:
                            try:
                                val = float(cv)
                                tgt_ws.Cells(r, vi + 3).Value = round(val, 2)
                                tgt_ws.Cells(r, vi + 3).NumberFormat = "0.00"
                                total_accum[vi] = total_accum.get(vi, 0) + val
                            except (ValueError, TypeError):
                                tgt_ws.Cells(r, vi + 3).Value = cv
                r += 1

            # Total row
            tgt_ws.Cells(r, 1).Value = "合计"
            tgt_ws.Cells(r, 1).Font.Bold = True
            for vi, fp_val in enumerate(canonical_fps):
                if vi in total_accum:
                    tgt_ws.Cells(r, vi + 3).Value = round(total_accum[vi], 2)
                    tgt_ws.Cells(r, vi + 3).NumberFormat = "0.00"
                tgt_ws.Cells(r, vi + 3).Font.Bold = True
            tgt_ws.Range(
                tgt_ws.Cells(r, 1), tgt_ws.Cells(r, virtual_cols)
            ).Interior.Color = 0xE8F0FE
            virtual_end_row = r

            # Format virtual table
            tgt_ws.Columns(1).ColumnWidth = 5
            max_unit_len = max((len(info.get("unit_name", "")) for info in items), default=0)
            tgt_ws.Columns(2).ColumnWidth = max(12, max_unit_len)
            for vi in range(3, virtual_cols + 1):
                _max_w = 0.0
                _v = tgt_ws.Cells(virtual_end_row, vi).Value
                if _v is not None and _is_numeric_com(_v):
                    _max_w = _com_estimate_width(_v)
                tgt_ws.Columns(vi).ColumnWidth = max(12, round(_max_w + 2))
            brd = tgt_ws.Range(
                tgt_ws.Cells(data_start_row - 3, 1),
                tgt_ws.Cells(virtual_end_row, virtual_cols),
            )
            brd.Borders.LineStyle = 1
            brd.Borders.Weight = 2
            brd.WrapText = True
            tgt_ws.Range(
                tgt_ws.Cells(data_start_row - 3, 2),
                tgt_ws.Cells(virtual_end_row, 2),
            ).HorizontalAlignment = 1
            tgt_ws.Range(
                tgt_ws.Cells(data_start_row - 3, 1),
                tgt_ws.Cells(virtual_end_row, virtual_cols),
            ).Font.Name = "宋体"
            tgt_ws.Range(
                tgt_ws.Cells(data_start_row - 3, 1),
                tgt_ws.Cells(virtual_end_row, virtual_cols),
            ).Font.Size = 9
            r += 2

            # ── Paste original worksheets (cross-workbook copy) ──
            current_row = r
            for info in items:
                try:
                    src_wb = app.Workbooks.Open(info["path"])
                    src_ws = src_wb.ActiveSheet
                    src_last_row = src_ws.UsedRange.Rows.Count
                    src_last_col = src_ws.UsedRange.Columns.Count

                    src_range = src_ws.Range(
                        src_ws.Cells(1, 1),
                        src_ws.Cells(src_last_row, src_last_col),
                    )
                    src_range.Copy()

                    tgt_ws.Activate()
                    tgt_cell = tgt_ws.Cells(current_row, 1)
                    tgt_ws.Paste(tgt_cell)
                    app.CutCopyMode = False

                    # 宋体 9pt for pasted data cells
                    _paste_range = tgt_ws.Range(
                        tgt_ws.Cells(current_row, 1),
                        tgt_ws.Cells(current_row + src_last_row - 1, src_last_col),
                    )
                    _paste_range.Font.Name = "宋体"
                    _paste_range.Font.Size = 9

                    # Find the 合计 row and use only its numeric values for column width
                    _total_row = None
                    for _pr in range(current_row, current_row + src_last_row):
                        if str(tgt_ws.Cells(_pr, 1).Value or '').strip() == '合计':
                            _total_row = _pr
                            break
                    if _total_row is not None:
                        for _pc in range(1, src_last_col + 1):
                            _pv = tgt_ws.Cells(_total_row, _pc).Value
                            if _pv is not None and _is_numeric_com(_pv):
                                _ew = _com_estimate_width(_pv)
                                _desired = round(_ew + 2)
                                _cur_w = tgt_ws.Columns(_pc).ColumnWidth or 0
                                if _desired > _cur_w:
                                    tgt_ws.Columns(_pc).ColumnWidth = _desired

                    # Right-align "制表人" cells; widen cols with signature prompts
                    _sig_kws = ["总经理签字", "分管领导审核", "财务审核", "业务审核"]
                    for _rr in range(current_row, current_row + src_last_row):
                        for _cc in range(1, src_last_col + 1):
                            _cell = tgt_ws.Cells(_rr, _cc)
                            _v = _cell.Value
                            if _v is None or not isinstance(_v, str):
                                continue
                            vs = str(_v)
                            if "制表人" in vs:
                                _cell.HorizontalAlignment = -4152
                                break
                            for _kw in _sig_kws:
                                if _kw in vs:
                                    _cw = _cell.EntireColumn.ColumnWidth
                                    if _cw < 20:
                                        _cell.EntireColumn.ColumnWidth = 20
                                    break

                    src_wb.Close(SaveChanges=False)
                    current_row += src_last_row + 3  # 3 blank rows between tables
                except Exception as e:
                    warnings_list.append(f"复制失败: {info['fname']} - {e}")

            # ── Page setup ──
            tgt_ws.PageSetup.Orientation = 2
            tgt_ws.PageSetup.PaperSize = 9
            tgt_ws.PageSetup.Zoom = False
            tgt_ws.PageSetup.FitToPagesWide = 1
            tgt_ws.PageSetup.FitToPagesTall = 9999
            LEFT_MARGIN_PT = round(left_margin_cm * 28.35, 2)
            RIGHT_MARGIN_PT = round(right_margin_cm * 28.35, 2)
            tgt_ws.PageSetup.LeftMargin = LEFT_MARGIN_PT
            tgt_ws.PageSetup.RightMargin = RIGHT_MARGIN_PT
            tgt_ws.PageSetup.PrintTitleRows = ""

            # ── Save (with dedup suffix for same-named excluded files) ──
            ym_part = f"_{ym_display}" if ym_display else ""
            disp_name = _disp(group_key)
            name_used = _name_counter.setdefault(disp_name, 0) + 1
            _name_counter[disp_name] = name_used
            safe_key = re.sub(r'[\\/:*?"<>|]', '_', disp_name)
            name_suffix = f"_{name_used}" if name_used > 1 else ""
            output_name = f"{safe_key}_工资表合集{ym_part}{name_suffix}.xlsx"
            output_dir_abs = str(Path(output_dir).resolve())
            Path(output_dir_abs).mkdir(parents=True, exist_ok=True)
            output_path = str(Path(output_dir_abs, output_name))
            if Path(output_path).exists():
                Path(output_path).unlink()  # WPS prompts on overwrite
            try:
                tgt_wb.SaveAs(output_path)
            except Exception as save_err:
                fallback_path = str(Path(output_dir_abs, f"合并工资表合集_{group_idx}.xlsx"))
                if Path(fallback_path).exists():
                    Path(fallback_path).unlink()
                tgt_wb.SaveAs(fallback_path)
                output_path = fallback_path
                warnings_list.append(f"文件名包含特殊字符，已保存为: {Path(fallback_path).name}")
            tgt_wb.Close(SaveChanges=True)
            output_files.append(output_path)

        if progress_callback:
            progress_callback(
                len(groups), len(groups),
                f"完成！生成 {len(output_files)} 个合并工资表文件",
            )

    except Exception as e:
        warnings_list.append(f"WPS 合并失败：{e}")
        raise
    finally:
        if app:
            try:
                app.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

    return output_files, warnings_list, {"total_groups": len(groups)}


# ── Small helper to avoid pulling in os.walk at module level ──
def os_walk(top: str):
    """Minimal recursive directory walk (replacement for os.walk)."""
    import os
    for root, dirs, files in os.walk(top):
        yield root, dirs, files
