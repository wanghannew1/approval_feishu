"""Integration tests for payroll merger and printer."""
import openpyxl
from pathlib import Path
import pytest
from app.payroll_merger import (
    _load_mapping_rules, get_big_org,
    _extract_unit_from_sheet, _extract_unit_from_signed,
    _extract_yearmon_from_signed, _format_yearmons,
)


class TestMergeWorkflow:
    """Test the merge workflow end-to-end (without COM, just scanning + grouping)."""

    def test_scan_signed_files(self, tmp_path):
        """Create real xlsx files, verify scanning logic works."""
        units = ["test_unit_a", "test_unit_b", "test_unit_c"]
        files_created = []
        for i, u in enumerate(units):
            f = tmp_path / f"signed_{u}202606工资表.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=2, column=1, value=f"名称:{u}_dept")
            wb.save(f)
            files_created.append(f)
        # Verify files exist
        for f in files_created:
            assert f.exists()
        # Verify extraction works
        for i, u in enumerate(units):
            result = _extract_unit_from_sheet(str(files_created[i]))
            assert u in result

    def test_extract_real_xlsx(self, tmp_path):
        """Create xlsx with proper structure, extract unit name."""
        f = tmp_path / "test_real.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="单位名称：")
        ws.cell(row=2, column=2, value="吉林大学遥感学院")
        wb.save(f)
        result = _extract_unit_from_sheet(str(f))
        assert "吉林大学遥感学院" in result

    def test_extract_unit_from_signed_filename(self):
        result = _extract_unit_from_signed("signed_吉林大学电子学院202606工资表.xlsx")
        assert result == "吉林大学电子学院"

    def test_extract_yearmon_from_signed_filename(self):
        result = _extract_yearmon_from_signed("signed_吉林大学电子学院202606工资表.xlsx")
        assert result == "202606"

    def test_format_yearmons_range(self):
        result = _format_yearmons({"202606", "202607", "202608"})
        assert "202606" in result
        assert "202608" in result
        # Consecutive months are joined with '-'
        assert "-" in result


class TestBatchPrintFlow:
    """Test batch print logic (mocked)."""

    def test_batch_print_empty_list(self):
        """Call batch_print with empty list — WPS unavailable on non-Windows."""
        from app.payroll_merger import batch_print
        s, f, fl = batch_print([], None)
        assert s == 0
        assert f == 0
        # On non-Windows, WPS is unavailable so the fail list reports it
        assert "WPS不可用" in fl or fl == []


class TestMappingRules:
    """Test mapping rules work end-to-end with real file."""

    def test_mapping_rules_load_from_template(self):
        rules = _load_mapping_rules()
        assert len(rules) > 0
        # Verify structure
        first = rules[0]
        assert len(first) == 4  # (match_type, pattern, big_org, is_excluded)

    def test_get_big_org_returns_tuple(self):
        result = get_big_org("测试单位")
        assert isinstance(result, tuple)
        assert len(result) == 2
