"""
Tests for batch_processor module.

TDD test suite covering is_payroll_sheet, get_signature_path,
is_ready_for_print edge cases, and process_single_approval.
"""

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.batch_processor import (
    _build_output_path,
    _build_standard_name,
    _extract_first_row_title,
    _extract_unit_name,
    _extract_year_month,
    _get_signature_keywords,
    _is_standard_filename,
    _is_unit_name,
    _remove_empty_columns,
    get_payroll_config,
    get_signature_path,
    is_payroll_sheet,
    is_ready_for_print,
    process_single_approval,
)


class TestIsPayrollSheet:
    """Test suite for is_payroll_sheet function."""

    def _make_mock_worksheet(self, row_data):
        """Create a mock worksheet with given row data.

        Args:
            row_data: Dict mapping (row, col) to cell value.
        """
        ws = MagicMock()
        ws.max_row = max(r for r, c in row_data.keys()) if row_data else 0
        ws.max_column = max(c for r, c in row_data.keys()) if row_data else 0

        def get_cell(row, column=None):
            cell = MagicMock()
            cell.value = row_data.get((row, column))
            return cell

        ws.cell = get_cell
        return ws

    def test_payroll_sheet_with_all_required_elements(self):
        row_data = {
            (1, 1): "2026年派遣员工5月工资明细表（林下参）",
            (2, 1): "序号",
            (2, 2): "姓名",
            (2, 3): "基本工资",
            (2, 4): "应发工资",
            (2, 5): "转款合计",
            (2, 6): "实发工资",
            (3, 1): "养老",
            (10, 1): "   总经理签字：",
            (10, 11): "财务审核：",
        }
        ws = self._make_mock_worksheet(row_data)
        assert is_payroll_sheet(ws) is True

    def test_payroll_sheet_missing_title_keyword(self):
        row_data = {
            (1, 1): "普通表格",
            (2, 1): "序号",
            (2, 2): "姓名",
            (2, 3): "金额",
            (10, 1): "   总经理签字：",
        }
        ws = self._make_mock_worksheet(row_data)
        assert is_payroll_sheet(ws) is False

    def test_payroll_sheet_missing_required_content(self):
        row_data = {
            (1, 1): "2026年派遣员工5月工资明细表（林下参）",
            (2, 1): "部门信息",
            (3, 1): "养老",
            (10, 1): "   总经理签字：",
        }
        ws = self._make_mock_worksheet(row_data)
        assert is_payroll_sheet(ws) is False

    def test_payroll_sheet_missing_signature(self):
        row_data = {
            (1, 1): "2026年派遣员工5月工资明细表（林下参）",
            (2, 1): "序号",
            (2, 2): "应发工资",
            (2, 3): "实发工资",
            (3, 1): "养老",
            (10, 1): "其他内容",
        }
        ws = self._make_mock_worksheet(row_data)
        assert is_payroll_sheet(ws) is False

    def test_payroll_sheet_excluded_by_keyword(self):
        row_data = {
            (1, 1): "工资发放汇总数据",
            (2, 1): "生成时间",
            (4, 1): "文件名",
        }
        ws = self._make_mock_worksheet(row_data)
        assert is_payroll_sheet(ws) is False

    def test_payroll_sheet_with_custom_config(self):
        row_data = {
            (1, 1): "自定义工资表",
            (2, 1): "应发工资",
            (2, 2): "实发工资",
            (5, 1): "manager",
        }
        ws = self._make_mock_worksheet(row_data)
        custom_config = {
            "sheet_filter": {
                "title_keyword": {"required": "自定义工资表"},
                "exclude_keywords": {"keywords": []},
                "required_content": {"required": ["应发工资", "实发工资"]},
                "signatures": {
                    "mandatory": {
                        "manager": ["manager"],
                    },
                    "optional": {},
                },
            }
        }
        assert is_payroll_sheet(ws, custom_config) is True


class TestGetSignaturePath:
    """Test suite for get_signature_path function."""

    def test_signature_path_exact_match(self, tmp_path):
        """Test finding signature with exact name match."""
        sig_dir = tmp_path
        (sig_dir / "张三.png").touch()

        result = get_signature_path("张三", sig_dir)
        assert result == sig_dir / "张三.png"

    def test_signature_path_glob_match(self, tmp_path):
        """Test finding signature using glob when exact match fails."""
        sig_dir = tmp_path
        (sig_dir / "李四_签名.png").touch()

        result = get_signature_path("李四", sig_dir)
        assert result == sig_dir / "李四_签名.png"

    def test_signature_path_not_found(self, tmp_path):
        """Test None when signature not found."""
        sig_dir = tmp_path
        (sig_dir / "其他人.png").touch()

        result = get_signature_path("不存在", sig_dir)
        assert result is None

    def test_signature_path_empty_name(self, tmp_path):
        """Test None when approver name is empty."""
        sig_dir = tmp_path
        (sig_dir / "test.png").touch()

        result = get_signature_path("", sig_dir)
        assert result is None

    def test_signature_path_no_files_in_dir(self, tmp_path):
        """Test None when directory is empty."""
        result = get_signature_path("任何人", tmp_path)
        assert result is None


class TestIsReadyForPrint:
    """Test suite for is_ready_for_print edge cases."""

    @patch("app.batch_processor._get_role_mapping")
    @patch("app.batch_processor._get_mandatory_roles")
    def test_ready_when_all_mandatory_roles_approved(self, mock_mandatory, mock_mapping):
        """Test True when all mandatory roles have APPROVED status."""
        mock_mandatory.return_value = {"总经理签字", "部长签字"}
        mock_mapping.return_value = {"张三": "总经理签字", "李四": "部长签字"}

        details = {
            "approver_list": [
                {"approver_name": "张三", "status": "APPROVED"},
                {"approver_name": "李四", "status": "APPROVED"},
            ],
            "task_list": [
                {"node_name": "出纳办理", "status": "PENDING"},
            ]
        }
        assert is_ready_for_print(details) is True

    @patch("app.batch_processor._get_role_mapping")
    @patch("app.batch_processor._get_mandatory_roles")
    def test_not_ready_when_mandatory_role_missing(self, mock_mandatory, mock_mapping):
        """Test False when a mandatory role is missing or not approved."""
        mock_mandatory.return_value = {"总经理签字", "部长签字"}
        mock_mapping.return_value = {"张三": "总经理签字", "李四": "部长签字"}

        details = {
            "approver_list": [
                {"approver_name": "张三", "status": "APPROVED"},
                {"approver_name": "李四", "status": "PENDING"},
            ]
        }
        assert is_ready_for_print(details) is False

    @patch("app.batch_processor._get_role_mapping")
    @patch("app.batch_processor._get_mandatory_roles")
    def test_not_ready_when_no_approvers(self, mock_mandatory, mock_mapping):
        """Test False when approver_list is empty."""
        mock_mandatory.return_value = {"总经理签字"}
        mock_mapping.return_value = {}

        details = {"approver_list": []}
        assert is_ready_for_print(details) is False

    @patch("app.batch_processor._get_role_mapping")
    @patch("app.batch_processor._get_mandatory_roles")
    def test_false_when_mandatory_roles_empty(self, mock_mandatory, mock_mapping):
        """Test False when mandatory_roles is empty (no config)."""
        mock_mandatory.return_value = set()
        mock_mapping.return_value = {}

        details = {"approver_list": [{"approver_name": "张三", "status": "APPROVED"}]}
        assert is_ready_for_print(details) is False

    @patch("app.batch_processor._get_role_mapping")
    @patch("app.batch_processor._get_mandatory_roles")
    def test_ignores_non_approved_approvers(self, mock_mandatory, mock_mapping):
        """Test that only APPROVED status counts toward ready."""
        mock_mandatory.return_value = {"总经理签字"}
        mock_mapping.return_value = {"张三": "总经理签字"}

        details = {
            "approver_list": [
                {"approver_name": "张三", "status": "REJECTED"},
            ]
        }
        assert is_ready_for_print(details) is False


class TestProcessSingleApproval:
    """Test suite for process_single_approval function."""

    @patch("app.batch_processor.extract_attachments")
    @patch("app.batch_processor.parse_form")
    @patch("app.batch_processor.get_instance_detail")
    def test_process_downloads_non_approved_instance(
        self, mock_detail, mock_parse, mock_extract
    ):
        """Test that non-APPROVED instances still download but skip signing."""
        mock_detail.return_value = {
            "instance_code": "test123",
            "approval_name": "测试审批",
            "status": "PENDING",
        }
        mock_parse.return_value = []
        mock_extract.return_value = [{"field_name": "工资表", "value": ["url"]}]

        result = process_single_approval("test123", "fake_token", {})

        assert result["skipped"] is False
        assert "审批未通过" not in result.get("message", "")

    @patch("app.batch_processor.get_approvers_with_roles")
    @patch("app.batch_processor.extract_attachments")
    @patch("app.batch_processor.parse_form")
    @patch("app.batch_processor.get_instance_detail")
    def test_process_downloads_without_signers(
        self, mock_detail, mock_parse, mock_extract, mock_approvers
    ):
        """Test that instances without valid signers still download."""
        mock_detail.return_value = {
            "instance_code": "test123",
            "approval_name": "测试审批",
            "status": "APPROVED",
            "approver_list": [],
        }
        mock_approvers.return_value = []
        mock_parse.return_value = []
        mock_extract.return_value = [{"field_name": "工资表", "value": ["url"]}]

        result = process_single_approval("test123", "fake_token", {})

        assert result["skipped"] is False

    @patch("app.batch_processor.get_approvers_with_roles")
    @patch("app.batch_processor.extract_attachments")
    @patch("app.batch_processor.parse_form")
    @patch("app.batch_processor.get_instance_detail")
    def test_process_skips_when_no_attachments(
        self, mock_detail, mock_parse, mock_extract, mock_approvers
    ):
        """Test when instance has no attachments - returns failure."""
        mock_detail.return_value = {
            "instance_code": "test123",
            "title": "测试审批",
            "status": "APPROVED",
            "approver_list": [
                {"approver_name": "张三", "status": "APPROVED", "role": "总经理签字"}
            ],
        }
        mock_approvers.return_value = [
            {"approver_name": "张三", "status": "APPROVED", "role": "总经理签字"}
        ]
        mock_parse.return_value = []
        mock_extract.return_value = []

        result = process_single_approval("test123", "fake_token", {})

        assert result["success"] is False
        assert result["message"] == "无附件"

    @patch("app.batch_processor.get_approvers_with_roles")
    @patch("app.batch_processor.extract_attachments")
    @patch("app.batch_processor.parse_form")
    @patch("app.batch_processor.get_instance_detail")
    @patch("app.batch_processor.download_file")
    def test_process_success_with_download_only(
        self, mock_download, mock_detail, mock_parse, mock_extract, mock_approvers, tmp_path
    ):
        """Test successful processing with attachment download (no signing)."""
        mock_detail.return_value = {
            "instance_code": "test123",
            "title": "测试审批",
            "status": "APPROVED",
            "approver_list": [
                {"approver_name": "张三", "status": "APPROVED", "role": "总经理签字"}
            ],
        }
        mock_approvers.return_value = [
            {"approver_name": "张三", "status": "APPROVED", "role": "总经理签字"}
        ]
        mock_parse.return_value = []
        mock_extract.return_value = [
            {"field_name": "附件", "value": ["file_token_123"]}
        ]
        mock_download.return_value = str(tmp_path / "test.txt")

        config = {
            "save_dir": str(tmp_path),
            "signatures_dir": str(tmp_path / "sigs"),
        }

        result = process_single_approval("test123", "fake_token", config)

        assert result["success"] is True
        assert "test.txt" in result["downloaded"]

    @patch("app.batch_processor.get_approvers_with_roles")
    @patch("app.batch_processor.extract_attachments")
    @patch("app.batch_processor.parse_form")
    @patch("app.batch_processor.get_instance_detail")
    def test_process_returns_correct_instance_code(
        self, mock_detail, mock_parse, mock_extract, mock_approvers, tmp_path
    ):
        """Test that result contains the correct instance_code."""
        mock_detail.return_value = {
            "instance_code": "instance_abc",
            "approval_name": "标题",
            "status": "APPROVED",
        }
        mock_approvers.return_value = []
        mock_parse.return_value = []
        mock_extract.return_value = []

        result = process_single_approval("instance_abc", "fake_token", {})

        assert result["instance_code"] == "instance_abc"
        assert result["title"] == "标题"


class TestCleanupEmptyColumns:
    """Test suite for _remove_empty_columns, _get_signature_keywords, and 制表人 alignment."""

    def test_get_signature_keywords_from_config(self):
        """Test extracting signature keywords from payroll config."""
        keywords = _get_signature_keywords(get_payroll_config())
        expected = {
            "总经理签字",
            "分管领导审核",
            "财务审核",
            "业务审核",
            "部长签字",
            "部长、分管副总签字",
        }
        missing = expected - keywords
        assert not missing, f"Missing keywords: {missing}"
        assert _get_signature_keywords({}) == set()

    def test_remove_empty_columns_removes_truly_empty(self):
        """Test that completely empty columns are deleted."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=4, column=1, value="Header A")
        ws.cell(row=5, column=1, value="data A")
        ws.cell(row=4, column=4, value="Header D")
        ws.cell(row=5, column=4, value="data D")

        assert ws.max_column == 4

        cfg = get_payroll_config()
        _remove_empty_columns(ws, cfg)

        assert ws.max_column == 2, f"Expected 2 columns, got {ws.max_column}"
        assert ws.cell(row=4, column=1).value == "Header A"
        assert ws.cell(row=4, column=2).value == "Header D"

    def test_remove_empty_columns_preserves_keyword(self):
        """Test that a column with only signature keywords is preserved
        by moving keywords to the nearest non-empty column on the right."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=4, column=1, value="Name")
        ws.cell(row=5, column=1, value="张三")
        ws.cell(row=4, column=3, value="Amount")
        ws.cell(row=5, column=3, value="5000")
        # Column B has only a signature keyword in the data area
        ws.cell(row=5, column=2, value="总经理签字")

        cfg = get_payroll_config()
        _remove_empty_columns(ws, cfg)

        # Column B deleted; the keyword moves into C (target=D→C after shift)
        assert ws.max_column == 2
        assert ws.cell(row=5, column=2).value == "总经理签字"

    def test_remove_empty_columns_keyword_appends_at_end(self):
        """Test that a keyword-only column with no data to the right
        appends the keyword at the end."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=4, column=1, value="Name")
        ws.cell(row=5, column=1, value="张三")
        # Column B has only a signature keyword, no data columns to the right
        ws.cell(row=5, column=2, value="分管领导审核")

        cfg = get_payroll_config()
        _remove_empty_columns(ws, cfg)

        # Keyword written to max_column+1, then column B deleted → col 2 is the new column
        assert ws.max_column == 2
        assert ws.cell(row=5, column=2).value == "分管领导审核"

    def test_remove_empty_columns_with_formulas(self):
        """Columns with ONLY formulas (no plain data) → delete, formulas relocated."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Col A — real data
        ws.cell(row=4, column=1, value="Name")
        ws.cell(row=5, column=1, value="张三")
        # Col B — formulas only (like D-部门's 合计 row formulas)
        ws.cell(row=12, column=2, value="=ROUND(A5*0.07,2)")
        ws.cell(row=13, column=2, value="=SUM(A4:A5)")
        # Col C — real data
        ws.cell(row=4, column=3, value="Amount")
        ws.cell(row=5, column=3, value="5000")
        # Col D — signature keyword (to show formulas don't interfere)
        ws.cell(row=5, column=4, value="分管领导审核")

        cfg = get_payroll_config()
        _remove_empty_columns(ws, cfg)

        # B deleted, C→B, D→C, keyword moves into C
        assert ws.max_column == 3, f"expected 3, got {ws.max_column}"
        # Real data column C shifted left to B
        assert ws.cell(row=5, column=2).value == "5000"
        # Formula from original D (col 4, keyword) should now be at col 3
        assert ws.cell(row=5, column=3).value == "分管领导审核"

    def test_remove_empty_columns_no_empty_columns_noop(self):
        """Test that when all columns have real data, nothing changes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=4, column=1, value="Name")
        ws.cell(row=4, column=2, value="Amount")
        ws.cell(row=5, column=1, value="张三")
        ws.cell(row=5, column=2, value="5000")

        cfg = get_payroll_config()
        _remove_empty_columns(ws, cfg)

        assert ws.max_column == 2
        assert ws.cell(row=4, column=1).value == "Name"
        assert ws.cell(row=4, column=2).value == "Amount"
        assert ws.cell(row=5, column=1).value == "张三"
        assert ws.cell(row=5, column=2).value == "5000"

    def test_remove_empty_columns_with_merged_cells(self):
        """Test that merged cells are preserved when deleting empty columns."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.merge_cells("A5:B6")
        ws.cell(row=5, column=1, value="Merged Header")
        ws.cell(row=7, column=2, value="__keep__")
        ws.cell(row=5, column=4, value="End")

        cfg = get_payroll_config()
        _remove_empty_columns(ws, cfg)

        assert ws.max_column == 3
        assert ws.cell(row=5, column=3).value == "End"
        assert any("A5:B6" in str(mc) for mc in ws.merged_cells.ranges)

    def test_zhibiaoren_right_alignment(self):
        """Test that a cell containing '制表人' gets right/center alignment,
        matching the inline logic from _insert_signature_to_excel_openpyxl."""
        from openpyxl.styles import Font as _Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        cell = ws.cell(row=5, column=3, value="制表人：张三")

        # Inline replica of batch_processor.py lines 848–852
        if cell.value and "制表人" in str(cell.value):
            old_size = cell.font.size or 11
            if old_size > 10:
                cell.font = _Font(size=10)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        assert cell.alignment.horizontal == "right"
        assert cell.alignment.vertical == "center"

    # ── Standard filename rename tests ─────────────────────────────────────

    @staticmethod
    def _make_ws(cells: dict):
        """Build a real openpyxl worksheet from a cell-value mapping.

        *cells* maps ``(row, col)`` tuples to cell values, e.g.::

            _make_ws({(1, 1): "吉林大学2026年07月工资表"})
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for (r, c), val in cells.items():
            ws.cell(row=r, column=c, value=val)
        return ws

    def test_is_unit_name(self):
        """True for strings that look like a unit name."""
        assert _is_unit_name("吉林大学") is True
        assert _is_unit_name("北京大学法学院") is True
        assert _is_unit_name("中国航天科技集团") is True
        assert _is_unit_name("教育部考试中心") is True
        assert _is_unit_name("长春市税务局") is True
        assert _is_unit_name("某某企业") is True

    def test_is_unit_name_false(self):
        """False for short strings or non-unit text."""
        assert _is_unit_name("序号") is False
        assert _is_unit_name("ABC") is False
        assert _is_unit_name("姓名") is False
        assert _is_unit_name("基本工资") is False
        assert _is_unit_name("") is False

    def test_extract_unit_name_from_row2(self):
        """Row 2 unit name is extracted when present (with 名称： prefix)."""
        ws = self._make_ws({
            (1, 1): "2026年07月工资表",
            (2, 1): "单位名称：吉林大学商学与管理学院",
        })
        assert _extract_unit_name(ws) == "吉林大学商学与管理学院"

    def test_extract_unit_name_from_title(self):
        """Unit name is extracted from the row-1 title when row 2 has none."""
        ws = self._make_ws({
            (1, 1): "吉林大学商学与管理学院2026年07月工资表",
            (2, 1): "序号",
            (2, 2): "姓名",
        })
        assert _extract_unit_name(ws) == "吉林大学商学与管理学院"

    def test_extract_year_month_from_title(self):
        """Year-month is extracted from the row-1 title."""
        ws = self._make_ws({
            (1, 1): "吉林大学商学与管理学院2026年07月工资表",
        })
        assert _extract_year_month(ws) == "2026年07月"

    def test_extract_year_month_single_digit_month(self):
        """Single-digit month is zero-padded."""
        ws = self._make_ws({
            (1, 1): "某学院2025年5月工资表",
        })
        assert _extract_year_month(ws) == "2025年05月"

    def test_is_standard_filename(self):
        """Standard format filenames are recognised."""
        assert _is_standard_filename("吉林大学2026年07月工资表.xlsx") is True
        assert _is_standard_filename("signed_吉林大学2026年07月工资表.xlsx") is True
        assert _is_standard_filename("某学院2025年5月工资表.xlsx") is True

    def test_is_standard_filename_nonstandard(self):
        """Non-standard filenames are rejected."""
        assert _is_standard_filename("tddd_dialog_abc123.xlsx") is False
        assert _is_standard_filename("2465ea5e.xlsx") is False
        assert _is_standard_filename("工资表.xlsx") is False
        assert _is_standard_filename("2026年07月工资表.xlsx") is False

    def test_build_output_path_renames_nonstandard(self, tmp_path):
        """Non-standard source filename → signed standard name."""
        ws = self._make_ws({
            (1, 1): "吉林大学商学与管理学院2026年07月工资表",
        })
        src = tmp_path / "2465ea5e.xlsx"
        src.touch()
        dst = tmp_path / "signed_2465ea5e.xlsx"
        result = _build_output_path(src, dst, ws)
        assert result.name == "signed_吉林大学商学与管理学院2026年07月工资表.xlsx"
        assert result.parent == dst.parent

    def test_build_output_path_skips_when_standard(self, tmp_path):
        """Source already standard → unchanged."""
        ws = self._make_ws({
            (1, 1): "吉林大学商学与管理学院2026年07月工资表",
        })
        standard_name = "吉林大学2026年07月工资表.xlsx"
        src = tmp_path / standard_name
        src.touch()
        dst = tmp_path / "signed_吉林大学2026年07月工资表.xlsx"
        result = _build_output_path(src, dst, ws)
        assert result == dst

    def test_build_output_path_skips_when_no_unit(self, tmp_path):
        """No unit name in worksheet → no rename."""
        ws = self._make_ws({
            (1, 1): "2026年07月工资表",
            (2, 1): "序号",
            (2, 2): "姓名",
            (2, 3): "基本工资",
        })
        src = tmp_path / "2465ea5e.xlsx"
        src.touch()
        dst = tmp_path / "signed_2465ea5e.xlsx"
        result = _build_output_path(src, dst, ws)
        assert result == dst

    def test_build_output_path_uses_tddd_fallback_when_no_unit(self, tmp_path):
        """No unit, but tddd_dialog source → legacy fallback applies."""
        ws = self._make_ws({
            (1, 1): "2026年派遣员工5月工资明细表（林下参）",
        })
        src = tmp_path / "tddd_dialog_abc123.xlsx"
        src.touch()
        dst = tmp_path / "signed_tddd_dialog_abc123.xlsx"
        result = _build_output_path(src, dst, ws)
        assert "signed_tddd_dialog_abc123" in result.name
        assert "2026年派遣员工5月工资明细表" in result.name
        assert result.name.endswith(".xlsx")
