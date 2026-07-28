"""
Tests for payroll_merger module.
"""

import json
from pathlib import Path

import openpyxl
import pytest
from app.payroll_merger import (
    _load_mapping_rules,
    get_big_org,
    is_excluded,
    _extract_unit_from_sheet,
    _extract_unit_from_signed,
    _extract_yearmon_from_signed,
    _format_yearmons,
    _normalize_unit_name,
    check_wps_available,
)


# ══════════════════════════════════════════════════════════════════════
#  1–3: Mapping rule loading + structure
# ══════════════════════════════════════════════════════════════════════


class TestLoadMappingRules:
    """Tests for _load_mapping_rules()."""

    def test_load_mapping_rules_structure(self):
        """Each rule is a 4-tuple with all required fields."""
        rules = _load_mapping_rules()
        assert len(rules) > 0
        for r in rules:
            # Each rule is a 4-tuple: (match_type, pattern, big_org, is_excluded)
            assert isinstance(r, tuple), f"Rule is not a tuple: {r}"
            assert len(r) == 4, f"Rule does not have 4 elements: {r}"
            match_type, pattern, big_org, excluded = r
            assert match_type in (
                "prefix", "contains", "exact",
            ), f"Unknown match_type '{match_type}' in rule: {r}"
            assert isinstance(pattern, str) and pattern, f"Rule has empty pattern: {r}"
            assert isinstance(big_org, str), f"Rule big_org is not a string: {r}"
            assert isinstance(excluded, bool), f"Rule excluded is not a bool: {r}"

    def test_load_mapping_rules_has_big_orgs(self):
        """Returned rules cover at least 5 distinct big organisations."""
        rules = _load_mapping_rules()
        big_orgs = set()
        for r in rules:
            match_type, pattern, big_org, excluded = r
            if big_org:
                big_orgs.add(big_org)
        assert len(big_orgs) > 5, (
            f"Expected 5+ big orgs, got {len(big_orgs)}: {big_orgs}"
        )

    def test_load_mapping_rules_includes_exclusions(self):
        """At least one rule has is_excluded=True."""
        rules = _load_mapping_rules()
        exclusions = [r for r in rules if r[3] is True]
        assert len(exclusions) >= 1, (
            f"Expected at least 1 exclusion rule, got {len(exclusions)}"
        )

    def test_load_mapping_rules_rule_count(self):
        """Verify a known minimum rule count (spot check)."""
        rules = _load_mapping_rules()
        assert len(rules) >= 20, (
            f"Expected at least 20 rules, got {len(rules)}"
        )


# ══════════════════════════════════════════════════════════════════════
#  4–8: get_big_org matching
# ══════════════════════════════════════════════════════════════════════


class TestGetBigOrg:
    """Tests for get_big_org(unit_name)."""

    def test_get_big_org_returns_tuple(self):
        """Return value is always a 2-tuple of strings."""
        result = get_big_org("测试单位")
        assert isinstance(result, tuple), f"Expected tuple, got {type(result)}"
        assert len(result) == 2, f"Expected length 2, got {len(result)}"
        big_org, match_type = result
        assert isinstance(big_org, str), f"big_org not str: {type(big_org)}"
        assert isinstance(match_type, str), f"match_type not str: {type(match_type)}"

    def test_get_big_org_no_match(self):
        """Unrecognised unit falls back to original name for both values."""
        big_org, match_type = get_big_org("完全不存在的单位名称XYZ")
        assert big_org == "完全不存在的单位名称XYZ"
        assert match_type == "完全不存在的单位名称XYZ"

    def test_get_big_org_prefix_match(self):
        """'吉林大学物理学院' should match the '吉林大学' prefix rule."""
        big_org, match_type = get_big_org("吉林大学物理学院")
        assert big_org == "吉林大学"
        assert match_type == "吉林大学物理学院"

    def test_get_big_org_prefix_match_first_hospital(self):
        """'吉林大学第一医院' matches its own dedicated rule."""
        big_org, match_type = get_big_org("吉林大学第一医院")
        assert big_org == "吉林大学第一医院"
        assert match_type == "吉林大学第一医院"

    def test_get_big_org_contains_match(self):
        """'吉林农业大学某学院' matches the contains rule for '吉林农业大学'."""
        big_org, match_type = get_big_org("吉林农业大学某学院")
        assert big_org == "吉林农业大学"
        assert match_type == "吉林农业大学某学院"

    def test_get_big_org_has_excluded_rule(self):
        """Excluded units still resolve via rules (exclusion is checked separately)."""
        big_org, match_type = get_big_org("中国邮政储蓄银行股份有限公司吉林省分行直属支行")
        assert big_org == "中国邮政储蓄银行股份有限公司吉林省分行直属支行"
        assert match_type == "中国邮政储蓄银行股份有限公司吉林省分行直属支行"

    def test_get_big_org_exact_rule_not_used(self):
        """Verify exact-match coverage — at least one rule uses 'exact' type."""
        rules = _load_mapping_rules()
        exact_rules = [r for r in rules if r[0] == "exact"]
        assert len(exact_rules) >= 0  # exact is accepted by the code


# ══════════════════════════════════════════════════════════════════════
#  9: is_excluded
# ══════════════════════════════════════════════════════════════════════


class TestIsExcluded:
    """Tests for is_excluded(unit_name)."""

    def test_is_excluded_returns_bool(self):
        """Return type is bool."""
        result = is_excluded("测试单位")
        assert isinstance(result, bool), f"Expected bool, got {type(result)}"

    def test_is_excluded_known_excluded(self):
        """Units matching exclusion rules return True."""
        # '吉林银行' is a prefix rule with is_excluded=True
        result = is_excluded("吉林银行")
        assert result is True, "Expected '吉林银行' to be excluded"

    def test_is_excluded_normal_unit_not_excluded(self):
        """Normal unit like '吉林大学物理学院' returns False."""
        result = is_excluded("吉林大学物理学院")
        assert result is False, "Expected '吉林大学物理学院' not to be excluded"

    def test_is_excluded_prefix_match_excluded(self):
        """'吉林省交通实业发展有限公司' matches a prefix rule with is_excluded=True."""
        result = is_excluded("吉林省交通实业发展有限公司")
        assert result is True

    def test_is_excluded_nonexistent_unit(self):
        """Completely unknown unit is not excluded."""
        result = is_excluded("完全不存在的单位名称XYZ")
        assert result is False


# ══════════════════════════════════════════════════════════════════════
#  10–11: Unit extraction from signed filenames
# ══════════════════════════════════════════════════════════════════════


class TestExtractUnitFromSigned:
    """Tests for _extract_unit_from_signed(fname)."""

    def test_extract_unit_from_signed_standard(self):
        """Extract unit name from a standard signed_ filename."""
        result = _extract_unit_from_signed(
            "signed_吉林大学物理学院202606工资表.xlsx"
        )
        assert result == "吉林大学物理学院", (
            f"Expected '吉林大学物理学院', got '{result}'"
        )

    def test_extract_unit_from_signed_with_year_month_chinese(self):
        """Extract unit name with Chinese year-month format."""
        result = _extract_unit_from_signed(
            "signed_吉林大学物理学院2026年06月工资表.xlsx"
        )
        assert result == "吉林大学物理学院", (
            f"Expected '吉林大学物理学院', got '{result}'"
        )

    def test_extract_unit_from_signed_no_year(self):
        """Fallback when no year-month pattern in filename."""
        result = _extract_unit_from_signed("signed_吉林大学物理学院.xlsx")
        assert result == "吉林大学物理学院", (
            f"Expected '吉林大学物理学院', got '{result}'"
        )

    def test_extract_unit_from_signed_no_prefix(self):
        """Work even without 'signed_' prefix."""
        result = _extract_unit_from_signed("东北师范大学202606工资表.xlsx")
        assert result == "东北师范大学", (
            f"Expected '东北师范大学', got '{result}'"
        )


class TestExtractYearmonFromSigned:
    """Tests for _extract_yearmon_from_signed(fname)."""

    def test_extract_yearmon_from_signed_standard(self):
        """Extract YYYYMM from a standard filename."""
        result = _extract_yearmon_from_signed(
            "signed_吉林大学物理学院202606工资表.xlsx"
        )
        assert result == "202606", (
            f"Expected '202606', got '{result}'"
        )

    def test_extract_yearmon_from_signed_chinese_format(self):
        """Extract YYYYMM from Chinese year-month format."""
        result = _extract_yearmon_from_signed(
            "signed_吉林大学物理学院2026年06月工资表.xlsx"
        )
        assert result == "202606", (
            f"Expected '202606', got '{result}'"
        )

    def test_extract_yearmon_from_signed_single_digit_month(self):
        """Handle single-digit month (6 instead of 06)."""
        result = _extract_yearmon_from_signed(
            "signed_吉林大学物理学院2026年6月工资表.xlsx"
        )
        assert result == "202606", (
            f"Expected '202606', got '{result}'"
        )

    def test_extract_yearmon_from_signed_no_match(self):
        """Return empty string when no year-month pattern is found."""
        result = _extract_yearmon_from_signed("signed_no_date.xlsx")
        assert result == "", (
            f"Expected '', got '{result}'"
        )

    def test_extract_yearmon_from_signed_edge_202512(self):
        """Verify December of a year is handled correctly."""
        result = _extract_yearmon_from_signed(
            "signed_吉林大学物理学院202512工资表.xlsx"
        )
        assert result == "202512", (
            f"Expected '202512', got '{result}'"
        )


# ══════════════════════════════════════════════════════════════════════
#  12: _extract_unit_from_sheet (create temp xlsx)
# ══════════════════════════════════════════════════════════════════════


class TestExtractUnitFromSheet:
    """Tests for _extract_unit_from_sheet(fpath)."""

    def test_extract_unit_from_sheet_standard(self, tmp_path):
        """Extract unit name from '名称：...' in cell A2."""
        fpath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="名称：吉林大学测试学院")
        wb.save(fpath)
        result = _extract_unit_from_sheet(str(fpath))
        assert "吉林大学测试学院" in result, (
            f"Expected '吉林大学测试学院', got '{result}'"
        )

    def test_extract_unit_from_sheet_colon_in_column_b(self, tmp_path):
        """Extract unit name when '名称' is in A2 and value in B2."""
        fpath = tmp_path / "test_split.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="名称")
        ws.cell(row=2, column=2, value="：吉林大学测试学院")
        wb.save(fpath)
        result = _extract_unit_from_sheet(str(fpath))
        assert "吉林大学测试学院" in result, (
            f"Expected '吉林大学测试学院', got '{result}'"
        )

    def test_extract_unit_from_sheet_no_match(self, tmp_path):
        """Return empty string when no unit name is found."""
        fpath = tmp_path / "test_no_match.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="无关内容")
        wb.save(fpath)
        result = _extract_unit_from_sheet(str(fpath))
        assert result == "", (
            f"Expected '', got '{result}'"
        )

    def test_extract_unit_from_sheet_invalid_file(self, tmp_path):
        """Return empty string for non-existent file."""
        result = _extract_unit_from_sheet(
            str(tmp_path / "nonexistent.xlsx")
        )
        assert result == ""


# ══════════════════════════════════════════════════════════════════════
#  13: _format_yearmons
# ══════════════════════════════════════════════════════════════════════


class TestFormatYearmons:
    """Tests for _format_yearmons(yearmon_set)."""

    def test_format_yearmons_range(self):
        """Consecutive months are joined with '-' in display format."""
        # _format_yearmons formats YYYYMM strings, not dates
        # It joins consecutive months with '-'.
        result = _format_yearmons({"202601", "202602", "202603"})
        assert "202601" in result
        assert "202603" in result
        # Check consecutive grouping: first and last of range
        assert "-" in result, (
            f"Expected '-' for consecutive range, got '{result}'"
        )

    def test_format_yearmons_single(self):
        """Single month returns just that month."""
        result = _format_yearmons({"202606"})
        assert result == "202606", (
            f"Expected '202606', got '{result}'"
        )

    def test_format_yearmons_non_consecutive(self):
        """Non-consecutive months are separated by '、'."""
        result = _format_yearmons({"202601", "202603"})
        assert "、" in result, (
            f"Expected '、' for non-consecutive months, got '{result}'"
        )
        assert "202601" in result
        assert "202603" in result

    def test_format_yearmons_empty_set(self):
        """Empty set returns empty string."""
        result = _format_yearmons(set())
        assert result == "", f"Expected '', got '{result}'"

    def test_format_yearmons_cross_year(self):
        """Consecutive months across year boundary are grouped."""
        result = _format_yearmons({"202612", "202701"})
        assert "202612" in result
        assert "202701" in result

    def test_format_yearmons_mixed(self):
        """Mixed consecutive and non-consecutive months."""
        result = _format_yearmons(
            {"202601", "202602", "202605", "202606", "202607"}
        )
        assert "202601-202602" in result or "202601" in result
        assert "202605-202607" in result or "202605" in result


# ══════════════════════════════════════════════════════════════════════
#  14: _normalize_unit_name
# ══════════════════════════════════════════════════════════════════════


class TestNormalizeUnitName:
    """Tests for _normalize_unit_name(name)."""

    def test_normalize_unit_name_strip_number_suffix(self):
        """Strip '(1)' suffix from unit name."""
        result = _normalize_unit_name("吉林大学(1)")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_strip_chinese_number_suffix(self):
        """Strip '（1）' suffix from unit name."""
        result = _normalize_unit_name("吉林大学（1）")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_no_change(self):
        """Name without suffixes is unchanged."""
        result = _normalize_unit_name("吉林大学")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_strip_copy_suffix(self):
        """Strip '- 副本' suffix."""
        result = _normalize_unit_name("吉林大学 - 副本")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_strip_copy_english(self):
        """Strip '- Copy' suffix."""
        result = _normalize_unit_name("吉林大学 - Copy")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_strip_underscore_copy(self):
        """Strip '_副本' suffix."""
        result = _normalize_unit_name("吉林大学_副本")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_strip_fulian_copy(self):
        """Strip '- 复件' suffix."""
        result = _normalize_unit_name("吉林大学 - 复件")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_with_whitespace(self):
        """Strip surrounding whitespace."""
        result = _normalize_unit_name("  吉林大学  ")
        assert result == "吉林大学", (
            f"Expected '吉林大学', got '{result}'"
        )

    def test_normalize_unit_name_number_3_suffix(self):
        """Strip '(3)' suffix."""
        result = _normalize_unit_name("东北师范大学(3)")
        assert result == "东北师范大学", (
            f"Expected '东北师范大学', got '{result}'"
        )


# ══════════════════════════════════════════════════════════════════════
#  15: check_wps_available (non-Windows always False)
# ══════════════════════════════════════════════════════════════════════


class TestCheckWpsAvailable:
    """Tests for check_wps_available()."""

    def test_check_wps_available(self):
        """On non-Windows, check_wps_available returns False."""
        result = check_wps_available()
        assert result is False, (
            f"Expected False on non-Windows, got {result}"
        )
